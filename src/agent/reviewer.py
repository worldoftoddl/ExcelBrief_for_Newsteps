"""조서검토 전용 그래프 — langgraph.json의 "reviewer" 진입점.

조서 완성도 점검을 고정 워크플로로 수행한다 (범용 해설은 agent 그래프 몫):

  triage(검토/대화 분기) → locate(파일 탐지) → collect(기계 증거 수집)
  → investigate(증거 충분성 점검·보충 조사, 미니 ReAct) → assess(LLM
  구조화 소견) → cite(기준서 근거 검색·확정, MCP) → report(템플릿 렌더)

인사·사용법·이전 보고서에 대한 후속 질문은 triage가 chat 노드로 보낸다
(LLM 분류 1회, 실패 시 파일 언급 유무 휴리스틱 폴백).

역할 분담: investigate만 도구 재량(Excel 도구, 호출 상한)을 갖고, assess는
순수 구조화 출력, cite는 LLM 없이 결정적으로 standards_search →
standards_get_paragraph 재확인을 수행한다 — 검색어 생성은 LLM(assess)이,
검색 실행·인용 확정은 코드가 맡는다. MCP 미연결 시 cite는 인용 없이
통과한다(우아한 강등). 보고서 형식은 템플릿이 고정한다.

한계(보고서에 고지): 기준서·지침 근거 인용은 이 그래프가 하지 않는다 —
근거가 필요하면 agent 그래프(기준서 도구 보유)를 사용한다.
"""

import json
from typing import Any, Literal

from langchain_core.messages import (
    AIMessage,
    HumanMessage,
    SystemMessage,
    ToolMessage,
)
from langchain_core.runnables import RunnableConfig
from langgraph.graph import END, START, StateGraph
from langgraph.graph.message import add_messages
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, field_validator
from typing_extensions import Annotated, TypedDict

from agent.graph import DEFAULT_MODEL, resolve_model
from agent.graph_common import (
    conversation_context,
    emit,
    find_target_file,
    human_texts_newest_first,
    missing_file_message,
)
from agent.mcp_client import get_standards_tools
from agent.tools.excel import (
    _detect_blocks,
    _load,
    _resolve,
    excel_find,
    excel_formula_map,
    excel_get_annotations,
    excel_read_range,
    excel_sheet_stats,
    excel_workbook_overview,
    list_workpapers,
)

MAX_SHEETS = 6
MAX_BLOCK_CELLS = 400  # 시트당 본문 정독 상한 (excel_read_range 상한 500 미만)
MAX_EVIDENCE_CHARS = 28_000
MAX_ASSESS_ATTEMPTS = 2

# investigate 미니 ReAct의 상한 — 재량은 주되 폭주는 구조로 막는다
MAX_INVESTIGATE_ROUNDS = 3  # 모델 응답 횟수
MAX_INVESTIGATE_CALLS = 6  # 도구 호출 총량
MAX_TOOL_RESULT_CHARS = 4_000  # 호출당 결과 클립
MAX_EXTRA_EVIDENCE_CHARS = 8_000  # 추가 증거 총량

MAX_CITED_FINDINGS = 5  # cite가 근거를 찾는 소견 수 (심각도순)

INVESTIGATE_TOOLS = (
    excel_read_range,
    excel_sheet_stats,
    excel_find,
    excel_formula_map,
    excel_get_annotations,
)

SIGNOFF_KEYWORDS = (
    "작성자", "검토자", "작성일", "검토일", "서명", "확인자",
    "Preparer", "Reviewer", "Prepared by", "Reviewed by",
)


# ── 서명란 스캔 (비LLM) ──────────────────────────────────────────────────
def _scan_signoffs(target) -> str:
    """작성/검토 표지 셀을 찾아 같은 셀·오른쪽·아래 셀의 채움 여부를 판정한다."""
    wb = _load(target, data_only=True)
    lines = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                text = value.strip()
                # 긴 문장 속 우연 일치(예: "…검토자는 다음을 확인한다")는 표지가 아님
                if len(text) > 30 or not any(k in text for k in SIGNOFF_KEYWORDS):
                    continue
                inline = text.split(":", 1)[1].strip() if ":" in text else ""
                right = ws.cell(row=cell.row, column=cell.column + 1).value
                below = ws.cell(row=cell.row + 1, column=cell.column).value
                filled = (
                    inline
                    or (str(right).strip() if right is not None else "")
                    or (str(below).strip() if below is not None else "")
                )
                status = f"채움({filled[:20]})" if filled else "공란"
                lines.append(f"- {ws.title}!{cell.coordinate} \"{text}\" → {status}")
    if not lines:
        return "[서명란 스캔] 작성·검토 표지를 찾지 못함"
    return "[서명란 스캔] (표지 셀 → 같은 셀·오른쪽·아래 값 유무)\n" + "\n".join(lines[:40])


def _clip_block_ref(block: dict) -> str:
    """블록을 정독 상한(MAX_BLOCK_CELLS) 이내로 행을 잘라 ref로 만든다."""
    rows = max(1, min(block["rows"], MAX_BLOCK_CELLS // max(1, block["cols"])))
    end_row = block["first_row"] + rows - 1
    return (
        f"{get_column_letter(block['c1'])}{block['first_row']}:"
        f"{get_column_letter(block['c2'])}{end_row}"
    )


def _collect_evidence(path_name: str) -> tuple[str, list[str], list[str]]:
    """검토에 필요한 기계 증거를 기존 도구 함수로 수집한다 (비LLM).

    (증거 텍스트, 점검한 시트, 생략한 시트)를 돌려준다 — 보고서가 점검
    범위를 결정적으로 표기할 수 있게. 서명란 스캔은 핵심 증거라 증거
    선두(개요 직후)에 둔다 — 뒤에 두면 MAX_EVIDENCE_CHARS 절단 시 가장
    먼저 잘려나간다.
    """
    target = _resolve(path_name)
    wb = _load(target, data_only=True)
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    examined = [ws.title for ws in visible[:MAX_SHEETS]]
    skipped = [ws.title for ws in visible[MAX_SHEETS:]]
    parts = [excel_workbook_overview.func(path_name), _scan_signoffs(target)]
    for ws in visible[:MAX_SHEETS]:
        parts.append(excel_formula_map.func(path_name, ws.title))
        parts.append(excel_get_annotations.func(path_name, ws.title))
        blocks = _detect_blocks(ws)
        if blocks:
            largest = max(blocks, key=lambda b: b["rows"] * b["cols"])
            parts.append(
                excel_read_range.func(path_name, ws.title, _clip_block_ref(largest))
            )
    if skipped:
        parts.append(
            f"(주의: 시트 {len(skipped)}개는 증거 수집에서 생략됨 — "
            f"상한 {MAX_SHEETS}개: {', '.join(skipped)})"
        )
    evidence = "\n\n".join(parts)
    if len(evidence) > MAX_EVIDENCE_CHARS:
        evidence = evidence[:MAX_EVIDENCE_CHARS] + "\n… (증거 절단 — 상한 초과)"
    return evidence, examined, skipped


# ── LLM 구조화 소견 ──────────────────────────────────────────────────────
class TriageDecision(BaseModel):
    mode: Literal["review", "chat"] = Field(
        description=(
            "review: 조서 파일의 완성도 점검·검토 요청. "
            "chat: 인사·사용법 질문·기능 문의·이전 검토 보고서에 대한 후속 "
            "설명 등 재검토가 필요 없는 대화."
        )
    )


class Finding(BaseModel):
    title: str = Field(description="소견 한 줄 요약")
    severity: Literal["높음", "중간", "낮음"]
    location: str = Field(description="근거 위치 — 시트!셀 또는 시트!범위, 없으면 시트명")
    detail: str = Field(description="무엇이 왜 문제/필요한지 신입 회계사가 이해할 설명")
    standards_query: str = Field(
        default="",
        description=(
            "이 소견의 근거가 될 기준서 검색어 — 한국어 핵심 개념 위주 "
            "(예: '수취채권 외부조회 확인절차'). 기준서 근거가 필요 없는 "
            "소견이면 빈 문자열"
        ),
    )
    source_hint: Literal["감사기준", "회계기준", "실무지침", ""] = Field(
        default="",
        description="검색을 한정할 문서군 — 불확실하면 빈 문자열(전체 검색)",
    )
    citation: str = Field(default="", description="(시스템이 채움) 확정 인용 표기")
    citation_cid: str = Field(default="", description="(시스템이 채움) 인용 cid")


class ReviewFindings(BaseModel):
    workpaper_purpose: str = Field(description="이 조서의 목적·대상 계정 한두 문장")
    performed_procedures: list[str] = Field(description="증거에서 확인되는 수행된 절차")
    missing_procedures: list[Finding] = Field(description="누락되었거나 미완성인 절차")
    signoff_assessment: str = Field(description="서명란 스캔 결과 해석 — 작성·검토 완료 여부")
    tieout_findings: list[Finding] = Field(
        description="검산·tie-out·하드코딩 관련 소견 (수식 지도의 하드코딩 숫자 포함)"
    )
    open_items: list[str] = Field(description="미결 항목·공란·추가 확인 필요 사항")
    overall: str = Field(description="완성도 총평 두세 문장")

    @field_validator("missing_procedures", "tieout_findings", mode="before")
    @classmethod
    def _coerce_str_findings(cls, value):
        """약한 모델이 Finding 객체 대신 문자열 리스트를 내는 경우를 승격한다 (Haiku 실측)."""
        if isinstance(value, list):
            return [
                {"title": v[:80], "severity": "중간", "location": "(미표기)", "detail": v}
                if isinstance(v, str)
                else v
                for v in value
            ]
        return value


async def _tool_text(tool, args: dict) -> str:
    """mcp_client 래퍼 도구를 직접 호출해 텍스트 콘텐츠만 뽑는다."""
    raw = await tool.coroutine(**args)
    content = raw[0] if isinstance(raw, tuple) else raw
    if isinstance(content, list):
        content = "\n".join(
            block.get("text", "") if isinstance(block, dict) else str(block)
            for block in content
        )
    return content if isinstance(content, str) else str(content)


def _first_hit(text: str) -> dict | None:
    """도구 결과 JSON에서 cid를 가진 첫 문단 아이템을 찾는다."""
    try:
        payload = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return None
    if not isinstance(payload, dict):
        return None
    if payload.get("cid"):
        return payload
    for key in ("results", "paragraphs"):
        items = payload.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            if isinstance(item, dict) and item.get("cid"):
                return item
    return None


class ReviewerState(TypedDict, total=False):
    messages: Annotated[list, add_messages]
    question: str
    mode: str  # "review" | "chat" — triage 분기 결과
    path: str
    evidence: str
    extra_evidence: str  # investigate가 보충한 증거
    examined: list  # 증거를 수집한 시트
    skipped: list  # 상한 초과로 생략된 시트
    findings: dict
    attempts: int
    error: str | None


def _render_report(
    path: str,
    findings: ReviewFindings,
    examined: list[str] | None = None,
    skipped: list[str] | None = None,
) -> str:
    def _finding_lines(items: list[Finding]) -> list[str]:
        if not items:
            return ["- (해당 없음)"]
        order = {"높음": 0, "중간": 1, "낮음": 2}
        lines = []
        for f in sorted(items, key=lambda f: order[f.severity]):
            line = f"- [{f.severity}] {f.title} ({f.location})\n  {f.detail}"
            if f.citation:
                line += f"\n  근거: {f.citation}"
            lines.append(line)
        return lines

    def _bullets(items: list[str]) -> list[str]:
        return [f"- {item}" for item in items] or ["- (해당 없음)"]

    lines = [
        f"# 조서 검토 보고 — {path}",
        "",
        "## ① 조서 개요",
        findings.workpaper_purpose,
        "",
        "## ② 수행된 절차 (증거 기준)",
        *_bullets(findings.performed_procedures),
        "",
        "## ③ 서명·검토 상태",
        findings.signoff_assessment,
        "",
        "## ④ 검산·tie-out 점검",
        *_finding_lines(findings.tieout_findings),
        "",
        "## ⑤ 누락·추가 필요 절차",
        *_finding_lines(findings.missing_procedures),
        "",
        "## ⑥ 미결 항목",
        *_bullets(findings.open_items),
        "",
        "## ⑦ 총평",
        findings.overall,
        "",
    ]
    if examined:
        lines += [
            "## 점검 범위",
            f"- 점검한 시트({len(examined)}개): {', '.join(examined)}",
            (
                f"- 생략된 시트({len(skipped)}개, 상한 {MAX_SHEETS}개 초과): "
                f"{', '.join(skipped)} — 필요하면 해당 시트만 담긴 파일로 다시 요청하세요."
                if skipped
                else "- 생략된 시트: 없음"
            ),
            "",
        ]
    cited = [
        f
        for f in [*findings.tieout_findings, *findings.missing_procedures]
        if f.citation_cid
    ]
    if cited:
        seen: set[str] = set()
        lines.append("## 근거 목록")
        for f in cited:
            if f.citation_cid in seen:
                continue
            seen.add(f.citation_cid)
            lines.append(f"- {f.citation} — `{f.citation_cid}`")
        lines.append("")
    lines += [
        "---",
        "*이 보고서는 조서 파일의 기계 수집 증거(구조·수식·주석·서명란)와 "
        "자동 보충 조사로 작성되었습니다. 기준서 근거는 소견별 자동 검색으로 "
        "원문을 재확인한 문단(cid 병기)만 인용하며, 인용이 없는 소견은 근거 "
        "미확정입니다. 심화 해석이 필요하면 agent 그래프(기준서 도구)를 "
        "사용하세요. 수식 값은 파일에 저장된 캐시 값 기준입니다.*",
    ]
    return "\n".join(lines)


class ReviewerNodes:
    def __init__(self, model) -> None:
        self.model = model

    def triage(self, state: ReviewerState) -> dict[str, Any]:
        """검토 요청인지 일반 대화인지 분기한다.

        LLM 구조화 출력 1회, 실패 시 휴리스틱(파일 언급 있으면 review,
        없으면 chat) 폴백. 이전 턴의 stale error도 여기서 리셋한다.
        """
        emit("triaging", "요청 유형을 판단하는 중")
        texts = human_texts_newest_first(state)
        question = texts[0].strip() if texts else ""
        if not question:
            return {"question": question, "mode": "chat", "error": None}
        has_file = find_target_file(texts) is not None
        context = conversation_context(state)
        context_part = f"이전 대화:\n{context}\n" if context else ""
        try:
            decider = self.model.with_structured_output(TriageDecision)
            result = decider.invoke(
                [
                    SystemMessage(
                        content=(
                            "당신은 감사조서 검토 에이전트의 라우터입니다. 사용자 "
                            "메시지가 조서 파일의 완성도 점검·검토 요청이면 review, "
                            "인사·사용법 질문·기능 문의·이전 검토 보고서에 대한 후속 "
                            "설명처럼 재검토가 필요 없는 대화면 chat으로 분류하세요. "
                            "다른 파일이나 같은 파일의 재검토를 새로 요청하면 "
                            "review입니다."
                        )
                    ),
                    HumanMessage(
                        content=(
                            f"{context_part}"
                            f"메시지: {question}\n"
                            f"메시지에 검토 대상 파일 언급 존재: {has_file}"
                        )
                    ),
                ]
            )
            decision = (
                result
                if isinstance(result, TriageDecision)
                else TriageDecision.model_validate(result)
            )
            mode = decision.mode
        except Exception:
            mode = "review" if has_file else "chat"
        return {"question": question, "mode": mode, "error": None}

    def route_triage(self, state: ReviewerState) -> Literal["review", "chat"]:
        return "chat" if state.get("mode") == "chat" else "review"

    def chat(self, state: ReviewerState) -> dict[str, Any]:
        """재검토 없이 답하는 일반 대화 — 역할·사용법·파일 목록을 컨텍스트로 준다."""
        emit("chatting", "일반 대화로 응답하는 중")
        response = self.model.invoke(
            [
                SystemMessage(
                    content=(
                        "당신은 '조서 검토 Agent'입니다 — 감사조서 파일의 완성도를 "
                        "기계 수집 증거(구조·수식·주석·서명란)로 점검해 보고서를 "
                        "만드는 검토 모드입니다. 지금 요청은 재검토가 필요 없는 일반 "
                        "대화로 분류되었습니다. 한국어로 간결히 답하고, 사용법을 "
                        "물으면 안내하세요: 검토할 파일을 첨부하거나 파일명을 언급해 "
                        "요청합니다. 이전 검토 보고서에 대한 후속 질문이면 대화 "
                        "맥락의 보고서 내용만으로 답하고, 증거를 새로 봐야 하는 "
                        "질문이면 재검토를 요청해 달라고 안내하세요. 기준서 번호 "
                        "인용은 하지 마세요 — 근거 인용은 agent 그래프(기준서 도구) "
                        "몫입니다.\n\n"
                        f"[현재 볼 수 있는 파일]\n{list_workpapers.func()}"
                    )
                ),
                *state["messages"],
            ]
        )
        emit("complete", "응답 완료")
        return {"messages": [response], "error": None}

    def locate(self, state: ReviewerState) -> dict[str, Any]:
        emit("locating", "검토할 조서 파일을 찾는 중")
        texts = human_texts_newest_first(state)
        question = texts[0].strip() if texts else ""
        target = find_target_file(texts)
        if target is None:
            return {
                "question": question,
                "error": missing_file_message("검토할 Excel 조서를 찾지 못했습니다."),
            }
        return {"question": question, "path": target.name, "attempts": 0, "error": None}

    def route_locate(self, state: ReviewerState) -> Literal["collect", "fail"]:
        return "fail" if state.get("error") else "collect"

    def collect(self, state: ReviewerState) -> dict[str, Any]:
        emit("collecting", f"증거 수집 중: {state['path']} (구조·수식·주석·서명란)")
        try:
            evidence, examined, skipped = _collect_evidence(state["path"])
            return {
                "evidence": evidence,
                "examined": examined,
                "skipped": skipped,
                "error": None,
            }
        except ValueError as exc:
            return {"error": str(exc)}

    def route_collect(self, state: ReviewerState) -> Literal["assess", "fail"]:
        return "fail" if state.get("error") else "assess"

    def investigate(self, state: ReviewerState) -> dict[str, Any]:
        """기본 증거의 충분성을 점검하고 부족분만 도구로 보충한다 (미니 ReAct).

        재량은 이 노드에만 허용하고 상한(라운드·호출 수·결과 길이)으로
        폭주를 막는다. 어떤 실패도 검토를 중단시키지 않는다 — 보충 없이
        기본 증거로 진행한다.
        """
        emit("investigating", "증거 충분성을 점검하고 부족분을 조사하는 중")
        path = state["path"]
        tools_by_name = {t.name: t for t in INVESTIGATE_TOOLS}
        try:
            model_with_tools = self.model.bind_tools(list(INVESTIGATE_TOOLS))
        except Exception:
            return {"extra_evidence": ""}

        messages: list = [
            SystemMessage(
                content=(
                    "당신은 감사조서 검토자입니다. 아래 기본 증거로 조서 완성도 "
                    "판단이 충분한지 점검하고, 부족한 부분만 도구로 보충 조사하세요 "
                    "— 예: 증거 수집에서 생략된 시트, 잘려 읽힌 블록의 나머지, "
                    "의심 가는 셀·수식 확인. 모든 도구의 path 인수는 "
                    f"'{path}'입니다. 도구 호출은 총 {MAX_INVESTIGATE_CALLS}회 "
                    "까지입니다. 증거가 이미 충분하면 도구를 호출하지 말고 "
                    "'증거 충분'이라고만 답하세요."
                )
            ),
            HumanMessage(
                content=(
                    f"검토 요청: {state.get('question', '조서 완성도를 점검해줘')}\n"
                    f"대상 파일: {path}\n\n[기본 증거]\n{state['evidence']}"
                )
            ),
        ]
        extra_parts: list[str] = []
        calls_used = 0
        for _ in range(MAX_INVESTIGATE_ROUNDS):
            try:
                response = model_with_tools.invoke(messages)
            except Exception:
                break
            tool_calls = getattr(response, "tool_calls", None) or []
            if not tool_calls:
                break
            messages.append(response)
            for call in tool_calls:
                if calls_used >= MAX_INVESTIGATE_CALLS:
                    messages.append(
                        ToolMessage(
                            content="(도구 호출 상한 도달)", tool_call_id=call["id"]
                        )
                    )
                    continue
                tool = tools_by_name.get(call["name"])
                try:
                    result = (
                        tool.func(**call["args"])
                        if tool
                        else f"오류: 알 수 없는 도구 {call['name']}"
                    )
                except Exception as exc:
                    result = f"오류: {exc}"
                calls_used += 1
                result = str(result)[:MAX_TOOL_RESULT_CHARS]
                extra_parts.append(f"[추가 조사: {call['name']} {call['args']}]\n{result}")
                messages.append(ToolMessage(content=result, tool_call_id=call["id"]))
        if calls_used:
            emit("investigating", f"보충 조사 {calls_used}건 수행")
        extra = "\n\n".join(extra_parts)[:MAX_EXTRA_EVIDENCE_CHARS]
        return {"extra_evidence": extra}

    def assess(self, state: ReviewerState) -> dict[str, Any]:
        attempt = state.get("attempts", 0) + 1
        emit("assessing", "수집된 증거로 완성도를 평가하는 중", attempt=attempt)
        assessor = self.model.with_structured_output(ReviewFindings)
        try:
            result = assessor.invoke(
                [
                    SystemMessage(
                        content=(
                            "당신은 회계법인의 감사조서 검토자입니다. 제공된 기계 수집 "
                            "증거(워크북 구조·수식 지도·주석·서명란 스캔)와 추가 조사 "
                            "결과만 사용해 조서 완성도를 평가하세요. 증거에 없는 내용을 "
                            "추정하지 말고, 모든 소견에 근거 위치(시트!셀)를 표기하세요. "
                            "수식 지도의 하드코딩 숫자는 오류일 수도 의도된 입력일 수도 "
                            "있으니 맥락으로 판단하되 단정하지 마세요. 기준서 번호를 "
                            "본문에 직접 인용하지 말고, 기준서 근거가 필요한 소견에는 "
                            "standards_query(한국어 검색어)와 source_hint를 채우세요 — "
                            "원문 확인과 인용 확정은 시스템이 수행합니다."
                        )
                    ),
                    HumanMessage(
                        content=(
                            f"검토 요청: {state.get('question', '조서 완성도를 점검해줘')}\n"
                            f"대상 파일: {state['path']}\n\n[수집 증거]\n{state['evidence']}"
                            + (
                                f"\n\n[추가 조사]\n{state['extra_evidence']}"
                                if state.get("extra_evidence")
                                else ""
                            )
                        )
                    ),
                ]
            )
            findings = (
                result
                if isinstance(result, ReviewFindings)
                else ReviewFindings.model_validate(result)
            )
            return {"findings": findings.model_dump(), "attempts": attempt, "error": None}
        except Exception as exc:  # 구조화 출력 실패·일시 오류 → 재시도 후 fail
            return {"attempts": attempt, "error": f"소견 생성 실패 — {exc}"}

    def route_assess(self, state: ReviewerState) -> Literal["report", "retry", "fail"]:
        if not state.get("error"):
            return "report"
        if state.get("attempts", 0) < MAX_ASSESS_ATTEMPTS:
            return "retry"
        return "fail"

    async def cite(self, state: ReviewerState) -> dict[str, Any]:
        """소견의 기준서 근거를 검색·재확인해 확정한다 (LLM 없음, 결정적).

        agent 그래프의 인용 규칙(search → get_paragraph 재확인)을 코드로
        강제한다. MCP 미연결·검색 실패·재확인 불일치는 해당 소견의 인용을
        생략할 뿐 검토를 중단시키지 않는다.
        """
        emit("citing", "소견의 기준서 근거를 검색·확정하는 중")
        findings = ReviewFindings.model_validate(state["findings"])
        tools = {t.name: t for t in await get_standards_tools()}
        search = tools.get("standards_search")
        get_para = tools.get("standards_get_paragraph")
        if search is None or get_para is None:
            return {"findings": findings.model_dump()}

        order = {"높음": 0, "중간": 1, "낮음": 2}
        targets = sorted(
            (
                f
                for f in [*findings.missing_procedures, *findings.tieout_findings]
                if f.standards_query
            ),
            key=lambda f: order[f.severity],
        )[:MAX_CITED_FINDINGS]
        cited = 0
        for finding in targets:
            try:
                args: dict[str, Any] = {"query": finding.standards_query, "top_k": 3}
                if finding.source_hint:
                    args["source_type"] = [finding.source_hint]
                hit = _first_hit(await _tool_text(search, args))
                if not hit:
                    continue
                cid = hit["cid"]
                confirmed = _first_hit(await _tool_text(get_para, {"cid": cid}))
                if not confirmed or confirmed.get("cid") != cid:
                    continue  # 원문 재확인 실패 — 인용을 남기지 않는다
                finding.citation = confirmed.get("display") or hit.get("display") or cid
                finding.citation_cid = cid
                cited += 1
            except Exception:
                continue
        if cited:
            emit("citing", f"기준서 근거 {cited}건 확정")
        return {"findings": findings.model_dump()}

    def report(self, state: ReviewerState) -> dict[str, Any]:
        emit("reporting", "검토 보고서를 작성하는 중")
        findings = ReviewFindings.model_validate(state["findings"])
        text = _render_report(
            state["path"],
            findings,
            state.get("examined"),
            state.get("skipped"),
        )
        emit("complete", "조서 검토 완료")
        return {"messages": [AIMessage(content=text)], "error": None}

    def fail(self, state: ReviewerState) -> dict[str, Any]:
        message = state.get("error") or "조서 검토에 실패했습니다."
        emit("failed", message)
        return {"messages": [AIMessage(content=f"오류: {message}")]}


async def reviewer(config: RunnableConfig):
    """요청 config로 모델을 정해 조서검토 그래프를 조립한다 (langgraph 서버가 호출)."""
    model_spec = (config.get("configurable") or {}).get("model", DEFAULT_MODEL)
    nodes = ReviewerNodes(resolve_model(model_spec))

    builder = StateGraph(ReviewerState)
    builder.add_node("triage", nodes.triage)
    builder.add_node("chat", nodes.chat)
    builder.add_node("locate", nodes.locate)
    builder.add_node("collect", nodes.collect)
    builder.add_node("investigate", nodes.investigate)
    builder.add_node("assess", nodes.assess)
    builder.add_node("cite", nodes.cite)
    builder.add_node("report", nodes.report)
    builder.add_node("fail", nodes.fail)

    builder.add_edge(START, "triage")
    builder.add_conditional_edges(
        "triage", nodes.route_triage, {"review": "locate", "chat": "chat"}
    )
    builder.add_edge("chat", END)
    builder.add_conditional_edges(
        "locate", nodes.route_locate, {"collect": "collect", "fail": "fail"}
    )
    builder.add_conditional_edges(
        "collect", nodes.route_collect, {"assess": "investigate", "fail": "fail"}
    )
    builder.add_edge("investigate", "assess")
    builder.add_conditional_edges(
        "assess",
        nodes.route_assess,
        {"report": "cite", "retry": "assess", "fail": "fail"},
    )
    builder.add_edge("cite", "report")
    builder.add_edge("report", END)
    builder.add_edge("fail", END)
    return builder.compile()
