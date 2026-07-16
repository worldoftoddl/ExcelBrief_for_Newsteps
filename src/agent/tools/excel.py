"""Excel 탐색 도구 v2 — 값·공간·논리에 서식·의도 채널을 더한다.

모든 도구는 WORKPAPERS_DIR 하위 파일만 접근하며(경로 탈출 차단),
실패는 예외 대신 "오류: …" 텍스트로 반환해 에이전트가 자가 수정하게 한다.
모든 도구 출력의 첫 줄은 출처(파일명 또는 시트!범위)다.

R1C1 정규화·블록 감지·이탈 휴리스틱은 docs/reference/xlsx_agent_tools.py에서
이식 (알고리즘만 — JSON 봉투·세션 상태는 ExcelBrief 방침에 따라 채택하지 않음).
"""

import inspect as _inspect
import os
import re
from collections import defaultdict
from functools import lru_cache
from pathlib import Path

import openpyxl.worksheet.dimensions as _dims
import xlrd
from langchain_core.tools import tool
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

# ── openpyxl 호환 패치 ────────────────────────────────────────────────
# 한공회 표준 서식 등 동아시아권 Excel의 열 속성(phonetic 등)이 openpyxl
# 3.1.5의 ColumnDimension이 모르는 kwarg로 들어와 로드 자체가 죽는다
# (실측: 감사조서서식_4000). 미지의 속성은 버리고 로드를 계속한다.
_orig_coldim_init = _dims.ColumnDimension.__init__
_COLDIM_PARAMS = set(_inspect.signature(_orig_coldim_init).parameters) - {"self"}


def _coldim_init_ignoring_unknown(self, worksheet, **kw):
    known = {k: v for k, v in kw.items() if k in _COLDIM_PARAMS}
    _orig_coldim_init(self, worksheet, **known)


_dims.ColumnDimension.__init__ = _coldim_init_ignoring_unknown

MAX_CELLS = 500
MAX_FIND_HITS = 50
MAX_BLOCKS_SHOWN = 8
MAX_PATTERNS_SHOWN = 20
MAX_DEVIANTS_SHOWN = 30


SUPPORTED_PATTERNS = ("*.xlsx", "*.xlsm", "*.xls", "*.docx")


def _base_dir() -> Path:
    return Path(os.environ.get("WORKPAPERS_DIR", "data/workpapers")).resolve()


def _supported_files(base: Path) -> list[Path]:
    files = {p for pat in SUPPORTED_PATTERNS for p in base.glob(pat)}
    return sorted(files, key=lambda p: p.name)


def _resolve(path: str) -> Path:
    """WORKPAPERS_DIR 하위로 경로를 한정한다. 실패 시 ValueError(사용자용 메시지)."""
    base = _base_dir()
    target = (base / path).resolve()
    if not target.is_relative_to(base):
        raise ValueError(f"오류: 조서 폴더({base.name}/) 밖의 경로는 접근할 수 없습니다.")
    if not target.is_file():
        candidates = [p.name for p in _supported_files(base)]
        listing = "\n".join(f"- {c}" for c in candidates) or "(폴더가 비어 있음)"
        raise ValueError(
            f"오류: '{path}' 파일이 없습니다. 현재 폴더의 파일 목록:\n{listing}"
        )
    return target


def _load_xls(path_str: str):
    """구형 .xls(BIFF)를 xlrd로 읽어 값 전용 openpyxl 워크북으로 변환한다.

    .xls 리더는 수식·서식·메모를 제공하지 않으므로 값·병합·숨김 시트만 옮긴다.
    """
    # formatting_info=True여야 merged_cells가 채워진다 (xlrd 2.x, .xls 전용)
    book = xlrd.open_workbook(path_str, formatting_info=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sh in book.sheets():
        ws = wb.create_sheet(title=sh.name[:31])
        if sh.visibility != 0:
            ws.sheet_state = "hidden"
        for r in range(sh.nrows):
            for col in range(sh.ncols):
                cell = sh.cell(r, col)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    v = xlrd.xldate.xldate_as_datetime(v, book.datemode)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    v = bool(v)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    v = xlrd.error_text_from_code.get(v, "#ERR")
                elif isinstance(v, float) and v.is_integer():
                    v = int(v)
                ws.cell(row=r + 1, column=col + 1, value=v)
        for rlo, rhi, clo, chi in sh.merged_cells:  # rhi·chi는 배타 경계
            ws.merge_cells(
                start_row=rlo + 1, end_row=rhi, start_column=clo + 1, end_column=chi
            )
    return wb


@lru_cache(maxsize=8)
def _load_cached(path_str: str, mtime: float, data_only: bool):
    if path_str.lower().endswith(".xls"):
        return _load_xls(path_str)
    return load_workbook(path_str, read_only=False, data_only=data_only)


def _load(target: Path, data_only: bool = True):
    """워크북 로드 — (경로, mtime, data_only) 키 LRU 캐시. 도구는 무상태 유지."""
    return _load_cached(str(target), target.stat().st_mtime, data_only)


def _sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise ValueError(
            f"오류: 시트 '{name}'이(가) 없습니다. 시트 목록: {', '.join(wb.sheetnames)}"
        )
    return wb[name]


def _fmt(value) -> str:
    return "" if value is None else str(value)


def _md(text: str) -> str:
    """마크다운 표 셀 안전화 — 파이프 이스케이프·개행 제거."""
    return text.replace("|", "\\|").replace("\n", " ")


def _color_code(color) -> str | None:
    """openpyxl 색 3계열(rgb/theme/indexed)을 압축 표기로. 해석은 에이전트 몫."""
    if color is None:
        return None
    kind = getattr(color, "type", None)
    if kind == "rgb":
        return color.rgb[-6:] if isinstance(color.rgb, str) else None
    if kind == "theme":
        tint = color.tint or 0.0
        return f"T{color.theme}" + (f"{tint:+.2f}" if abs(tint) > 1e-9 else "")
    if kind == "indexed":
        return f"I{color.indexed}"
    return None


def _format_marks(cell) -> list[str]:
    """셀 서식 압축 주석: B=볼드, F=배경색, C=글자색(기본 검정 계열 제외)."""
    marks = []
    if cell.font is not None and cell.font.bold:
        marks.append("B")
    if cell.fill is not None and cell.fill.patternType is not None:
        fill = _color_code(cell.fill.fgColor)
        if fill:
            marks.append(f"F:{fill}")
    if cell.font is not None:
        fc = _color_code(cell.font.color)
        if fc and fc not in ("000000", "T0", "T1"):
            marks.append(f"C:{fc}")
    return marks


# R1C1 정규화 — docs/reference/xlsx_agent_tools.py 이식.
# 한계: 문자열 리터럴 안의 유사 참조까지 치환될 수 있음 (구조 압축 목적이라 허용).
_REF_RE = re.compile(r"(?<![A-Z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)(?!\()")


def _to_r1c1(formula: str, base_row: int, base_col: int) -> str:
    def repl(m: re.Match) -> str:
        dc, col, dr, row = m.groups()
        ci, ri = column_index_from_string(col), int(row)
        c = f"C{ci}" if dc else ("C" if ci == base_col else f"C[{ci - base_col}]")
        r = f"R{ri}" if dr else ("R" if ri == base_row else f"R[{ri - base_row}]")
        return r + c

    return _REF_RE.sub(repl, formula)


def _detect_blocks(ws) -> list[dict]:
    """빈 행 2개 이상을 경계로 값-블록(테이블 후보)을 분할 — 참조 구현 이식."""
    row_span: dict[int, tuple[int, int]] = {}
    for row in ws.iter_rows():
        cells = [c for c in row if c.value is not None]
        if cells:
            row_span[cells[0].row] = (
                min(c.column for c in cells),
                max(c.column for c in cells),
            )
    groups, cur, prev = [], [], None
    for r in sorted(row_span):
        if prev is not None and r - prev > 2:
            groups.append(cur)
            cur = []
        cur.append(r)
        prev = r
    if cur:
        groups.append(cur)

    blocks = []
    for rows in groups:
        c1 = min(row_span[r][0] for r in rows)
        c2 = max(row_span[r][1] for r in rows)
        blocks.append(
            {
                "ref": f"{get_column_letter(c1)}{rows[0]}:{get_column_letter(c2)}{rows[-1]}",
                "rows": len(rows),
                "cols": c2 - c1 + 1,
                "first_row": rows[0],
                "c1": c1,
                "c2": c2,
            }
        )
    return blocks


@tool
def list_workpapers() -> str:
    """조서 폴더에 있는 파일 목록을 반환한다 (Excel: xlsx/xlsm/xls, Word: docx).

    파일명이 기억나지 않거나 사용자가 파일을 모호하게 지칭할 때,
    또는 사용자가 방금 업로드한 파일을 찾을 때 먼저 호출한다.
    """
    base = _base_dir()
    files = _supported_files(base)
    if not files:
        return f"조서 폴더({base})에 지원되는 파일(xlsx/xlsm/xls/docx)이 없습니다."
    return f"조서 폴더 파일 목록:\n" + "\n".join(
        f"- {p.name} ({p.stat().st_size // 1024:,} KB)" for p in files
    )


@tool
def excel_workbook_overview(path: str) -> str:
    """워크북 전체 지도: 시트별 크기·값/수식 밀도·시트간 참조·병합·값-블록 경계.

    어떤 Excel 파일이든 정독 전에 반드시 이 도구로 구조를 먼저 파악한다.
    블록 ref(예 "A4:H8")는 excel_read_range의 cell_range 인수로 그대로 넘긴다.

    Args:
        path: 조서 폴더 기준 파일명 (예: "감사조서서식_3650 감사 전 재무제표 확인.xlsx")
    """
    try:
        target = _resolve(path)
    except ValueError as e:
        return str(e)

    wb_v = _load(target, data_only=True)
    wb_f = _load(target, data_only=False)
    lines = [f"워크북: {target.name} — 시트 {len(wb_f.sheetnames)}개"]
    for ws in wb_f.worksheets:
        n_val = n_frm = n_x = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_frm += 1
                    if "!" in c.value:
                        n_x += 1
                else:
                    n_val += 1
        state = "" if ws.sheet_state == "visible" else f" ({ws.sheet_state})"
        lines.append(
            f"\n[{ws.title}]{state} {ws.max_row}행 × {ws.max_column}열, "
            f"병합 {len(ws.merged_cells.ranges)}건 — "
            f"값 {n_val:,}·수식 {n_frm:,}·시트간참조 {n_x:,}"
        )
        ws_v = wb_v[ws.title]
        blocks = _detect_blocks(ws)
        for i, b in enumerate(blocks[:MAX_BLOCKS_SHOWN], 1):
            preview_cells = []
            for col in range(b["c1"], min(b["c1"] + 8, b["c2"] + 1)):
                v = ws_v.cell(row=b["first_row"], column=col).value
                if v is not None:
                    preview_cells.append(str(v)[:20])
                if len(preview_cells) >= 6:
                    break
            preview = " | ".join(preview_cells) or "(값 없음)"
            lines.append(
                f"  블록{i} {b['ref']} ({b['rows']}행×{b['cols']}열) 첫 행: {preview}"
            )
        if len(blocks) > MAX_BLOCKS_SHOWN:
            lines.append(f"  … 블록 {len(blocks) - MAX_BLOCKS_SHOWN}개 생략")
    return "\n".join(lines)


@tool
def excel_read_range(path: str, sheet: str, cell_range: str, mode: str = "values") -> str:
    """지정 범위의 셀을 마크다운 표로 반환한다 (1회 최대 500셀).

    mode:
      - "values"(기본): 저장된 값 (수식은 파일에 저장된 캐시 값 — 재계산 아님)
      - "formulas": 수식 문자열 — 계산 로직 확인용
      - "format": 값 + 서식 주석 [B|F:색|C:색] — 색상 마킹·볼드 등 시각 단서 확인용.
        색은 RGB 6자리 또는 테마(T번호±틴트)·인덱스(I번호) 원값 — 의미(예: 노랑=검토완료)는
        조서 범례에서 파악할 것.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
        cell_range: 셀 범위 (예: "A1:C10") — overview의 블록 ref를 그대로 사용 가능
        mode: "values" | "formulas" | "format"
    """
    try:
        target = _resolve(path)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        count = (max_row - min_row + 1) * (max_col - min_col + 1)
        if count > MAX_CELLS:
            return (
                f"요청 범위가 {count:,}셀로 1회 상한({MAX_CELLS}셀)을 초과합니다. "
                f"범위를 더 작게 나눠 순차적으로 요청하세요."
            )
        wb = _load(target, data_only=(mode != "formulas"))
        ws = _sheet(wb, sheet)
    except ValueError as e:
        return str(e)

    header = [""] + [get_column_letter(c) for c in range(min_col, max_col + 1)]
    rows = ["| " + " | ".join(header) + " |",
            "|" + "---|" * len(header)]
    for r in range(min_row, max_row + 1):
        cells = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            text = _fmt(cell.value)
            if mode == "format":
                marks = _format_marks(cell)
                if marks:
                    text += f" [{'|'.join(marks)}]"
            cells.append(_md(text))
        rows.append("| " + " | ".join([str(r)] + cells) + " |")
    note = {
        "formulas": " (수식 문자열)",
        "format": " (값 [B=볼드|F=배경색|C=글자색; T=테마색, I=인덱스색])",
    }.get(mode, " (저장된 값 — 수식 재계산 아님)")
    if target.suffix.lower() == ".xls":
        note = " (.xls 구형 형식 — 수식·서식은 읽을 수 없어 저장된 값만 표시)"
    return f"{sheet}!{cell_range}{note}\n" + "\n".join(rows)


@tool
def excel_find(path: str, query: str, sheet: str = "", mode: str = "values") -> str:
    """문자열이 포함된 셀을 찾아 좌표와 값을 반환한다 (최대 50건).

    계정명·거래처·틱마크 등 특정 항목의 위치를 찾을 때 사용.
    mode="formulas"는 수식 문자열에서 검색 — 시트간 참조(예 "'1100'!") 추적용.

    Args:
        path: 조서 폴더 기준 파일명
        query: 찾을 문자열 (대소문자 무시, 부분 일치)
        sheet: 특정 시트로 한정할 때만 지정 (기본: 전체 시트)
        mode: "values"(기본) | "formulas"
    """
    try:
        target = _resolve(path)
        wb = _load(target, data_only=(mode != "formulas"))
        sheets = [_sheet(wb, sheet)] if sheet else wb.worksheets
    except ValueError as e:
        return str(e)

    needle = query.casefold()
    hits = []
    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and needle in str(cell.value).casefold():
                    hits.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
                    if len(hits) >= MAX_FIND_HITS:
                        return (
                            f"검색 '{query}' @ {target.name}\n" + "\n".join(hits)
                            + f"\n… 상한({MAX_FIND_HITS}건) 도달 — 검색어를 좁히세요."
                        )
    body = "\n".join(hits) if hits else "일치하는 셀이 없습니다."
    return f"검색 '{query}' @ {target.name}\n{body}"


@tool
def excel_get_annotations(path: str, sheet: str) -> str:
    """시트의 '적혀 있지 않은 의도'를 반환한다: 셀 메모(검토자 코멘트)·숨김
    행/열/시트·데이터 유효성(입력란 표지)·정의된 이름.

    조서에서 검토자가 남긴 코멘트, 숨겨진 전기 자료, 입력 셀 위치를 확인할 때 사용.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
    """
    try:
        target = _resolve(path)
        wb = _load(target, data_only=True)
        ws = _sheet(wb, sheet)
    except ValueError as e:
        return str(e)

    out = [f"주석·의도 정보: {target.name} [{sheet}]"]

    comments = [
        f"  {c.coordinate} ({c.comment.author}): {c.comment.text.strip()}"
        for row in ws.iter_rows()
        for c in row
        if c.comment is not None
    ]
    out.append(f"[메모] {len(comments)}건" if comments else "[메모] 없음")
    out += comments

    hidden_rows = [str(r) for r, d in sorted(ws.row_dimensions.items()) if d.hidden]
    hidden_cols = [c for c, d in sorted(ws.column_dimensions.items()) if d.hidden]
    hidden_sheets = [w.title for w in wb.worksheets if w.sheet_state != "visible"]
    out.append(
        f"[숨김] 행: {', '.join(hidden_rows) or '없음'} / "
        f"열: {', '.join(hidden_cols) or '없음'} / "
        f"시트: {', '.join(hidden_sheets) or '없음'}"
    )

    dvs = [
        f"  {dv.sqref} ({dv.type}: {dv.formula1})"
        for dv in ws.data_validations.dataValidation
    ]
    out.append(f"[데이터 유효성] {len(dvs)}건" if dvs else "[데이터 유효성] 없음")
    out += dvs

    dns = [f"  {name} → {dn.attr_text}" for name, dn in wb.defined_names.items()]
    out.append(f"[정의된 이름] {len(dns)}건" if dns else "[정의된 이름] 없음")
    out += dns

    return "\n".join(out)


@tool
def excel_formula_map(path: str, sheet: str) -> str:
    """시트 전체 수식을 R1C1 패턴으로 압축해 계산 골격(검산·tie-out·시트간 대사)을
    보여주고, 수식 지대 내 하드코딩 숫자(같은 행·열에 수식이 있는 자리의 상수)를
    검출한다.

    검출된 하드코딩이 오류인지 의도된 입력 셀인지는 조서 맥락으로 판단할 것 —
    이 도구는 검출만 한다.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
    """
    try:
        target = _resolve(path)
        ws = _sheet(_load(target, data_only=False), sheet)
    except ValueError as e:
        return str(e)

    if target.suffix.lower() == ".xls":
        return (
            f"수식 지도: {sheet} — .xls 구형 형식은 수식을 제공하지 않습니다. "
            f"값 확인은 excel_read_range를 사용하세요."
        )

    patterns: dict[str, list[str]] = defaultdict(list)
    frows, fcols = set(), set()
    numerics = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                patterns[_to_r1c1(v, c.row, c.column)].append(c.coordinate)
                frows.add(c.row)
                fcols.add(c.column)
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                numerics.append(c)

    if not patterns:
        return f"수식 지도: {sheet} — 수식 없음"

    total = sum(len(v) for v in patterns.values())
    lines = [f"수식 지도: {sheet} — 수식 {total}개, 패턴 {len(patterns)}종"]
    ranked = sorted(patterns.items(), key=lambda kv: -len(kv[1]))
    for i, (r1c1, coords) in enumerate(ranked[:MAX_PATTERNS_SHOWN], 1):
        lines.append(f"{i}. {r1c1[:100]} ×{len(coords)} (예: {', '.join(coords[:3])})")
    if len(ranked) > MAX_PATTERNS_SHOWN:
        lines.append(f"… 패턴 {len(ranked) - MAX_PATTERNS_SHOWN}종 생략")

    deviants = [
        f"  {c.coordinate}: {c.value}"
        for c in numerics
        if c.row in frows and c.column in fcols
    ]
    if deviants:
        lines.append(f"하드코딩 숫자(수식 지대 내): {len(deviants)}건")
        lines += deviants[:MAX_DEVIANTS_SHOWN]
        if len(deviants) > MAX_DEVIANTS_SHOWN:
            lines.append(f"  … {len(deviants) - MAX_DEVIANTS_SHOWN}건 생략")
    else:
        lines.append("하드코딩 숫자(수식 지대 내): 없음")
    return "\n".join(lines)


@tool
def excel_sheet_stats(path: str, sheet: str) -> str:
    """시트의 데이터 분포를 요약한다: 비어있지 않은 셀·수식 셀·숫자/문자 비율.

    큰 시트를 정독하기 전에 어느 영역부터 읽을지 우선순위를 정할 때 사용.

    Args:
        path: 조서 폴더 기준 파일명
        sheet: 시트 이름
    """
    try:
        target = _resolve(path)
        ws_f = _sheet(_load(target, data_only=False), sheet)
    except ValueError as e:
        return str(e)

    total = ws_f.max_row * ws_f.max_column
    non_empty = formulas = numbers = texts = 0
    val_r = val_c = 0
    for row in ws_f.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            non_empty += 1
            val_r, val_c = max(val_r, cell.row), max(val_c, cell.column)
            if isinstance(v, str) and v.startswith("="):
                formulas += 1
            elif isinstance(v, (int, float)):
                numbers += 1
            else:
                texts += 1
    density = non_empty / total * 100 if total else 0
    return (
        f"[{sheet}] {ws_f.max_row}행 × {ws_f.max_column}열 (총 {total:,}셀)\n"
        f"- 값 경계: {val_r}행 × {val_c}열 (서식만 있는 유령 행·열 제외)\n"
        f"- 비어있지 않은 셀: {non_empty:,} (밀도 {density:.0f}%)\n"
        f"- 수식 셀: {formulas:,}\n"
        f"- 숫자 셀: {numbers:,} / 문자 셀: {texts:,}\n"
        f"- 병합: {len(ws_f.merged_cells.ranges)}건"
    )


EXCEL_TOOLS = [
    list_workpapers,
    excel_workbook_overview,
    excel_read_range,
    excel_find,
    excel_get_annotations,
    excel_formula_map,
    excel_sheet_stats,
]
