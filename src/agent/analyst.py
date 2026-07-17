"""데이터 분석 전용 그래프 — langgraph.json의 "analyst" 진입점.

awesome-llm-apps/For_me/langgraph_data_analysis_agent의 고정 워크플로
(inspect → plan_sql → validate_sql → execute_sql → answer + revise 루프)를
채팅 UI용으로 각색해 이식했다. UI의 그래프 셀렉터에서 "agent"(범용 조서
해설, ReAct)와 별도로 선택하는 데이터 분석 모드다.

선두의 triage 노드가 요청을 분기한다 — 집계·계산 요청은 SQL 파이프라인
(inspect_data 이후)으로, 인사·사용법·후속 설명 등 일반 대화는 chat 노드로
보낸다 (LLM 분류 1회, 실패 시 파일 언급 유무 휴리스틱 폴백).

원본과 다른 점:
  - 입력이 {question}이 아니라 채팅 messages — 대상 파일은 메시지의
    "[첨부 파일: …]" 표기 또는 파일명 언급에서 찾는다
  - 데이터셋을 그래프 빌드 시 고정하지 않고, 요청마다 워크북에서 가장 큰
    값-블록(테이블 후보)을 골라 등록한다 (시트명이 언급되면 그 시트로 한정)
  - 등록·검증·실행은 tools/table.py의 DataStore 계층을 그대로 재사용
    (sqlglot AST 검증 + DuckDB enable_external_access=false, 2중 격리)
  - 모델은 메인 그래프와 동일하게 config["configurable"]["model"]로 라우팅
"""

import re
from typing import Any, Literal

from langchain_core.messages import AIMessage, HumanMessage, SystemMessage
from langchain_core.runnables import RunnableConfig
from langgraph.graph import END, START, StateGraph
from langgraph.graph.message import add_messages
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from pydantic import BaseModel, Field
from typing_extensions import Annotated, TypedDict

from agent.graph import DEFAULT_MODEL, resolve_model
from agent.graph_common import (
    conversation_context as _conversation_context,
    emit as _emit,
    find_target_file as _find_target_file,
    human_texts_newest_first as _human_texts_newest_first,
    missing_file_message,
)
from agent.tools.excel import _base_dir, _detect_blocks, _load, list_workpapers
from agent.tools.table import (
    MAX_RESULT_ROWS,
    TABLE,
    _markdown,
    _store_cached,
    _validate_sql,
)

MAX_SQL_REVISIONS = 2

# 질문 속 "시트명!A1:C50" 또는 "A1:C50" — 명시 범위는 자동 블록 선택보다 우선
_RANGE_RE = re.compile(
    r"(?:(?P<sheet>[\w.\-]{1,31})!)?"
    r"(?P<range>\$?[A-Za-z]{1,3}\$?\d{1,7}:\$?[A-Za-z]{1,3}\$?\d{1,7})"
)


class SQLPlan(BaseModel):
    sql: str = Field(description="One read-only DuckDB SELECT or WITH query")
    rationale: str


class TriageDecision(BaseModel):
    mode: Literal["analysis", "chat"] = Field(
        description=(
            "analysis: 표 데이터에 대한 집계·필터·정렬·계산 요청. "
            "chat: 인사·사용법 질문·기능 문의·이전 결과에 대한 후속 설명 등 "
            "SQL 실행이 필요 없는 대화."
        )
    )


class AnalystState(TypedDict, total=False):
    messages: Annotated[list, add_messages]
    question: str
    mode: str  # "analysis" | "chat" — triage 분기 결과
    dataset: dict  # {"path": 파일명, "sheet": 시트, "cell_range": 범위}
    schema: str
    profile: dict
    sql: str
    attempts: int
    error: str | None
    result_columns: list
    result_rows: list
    truncated: bool


def _trim_title_rows(ws, block: dict) -> str:
    """블록 선두의 제목·주석 행(비어있지 않은 셀 1개)을 잘라 헤더 행부터의 ref를 만든다.

    조서는 "제목 → 빈 행 1개 → 표" 배치가 흔한데 빈 행 1개는 블록 경계가
    되지 않아(_detect_blocks 기준: 2개 초과) 제목이 헤더로 오인된다.
    """
    _, _, _, max_row = range_boundaries(block["ref"])
    start = block["first_row"]
    for r in range(block["first_row"], max_row + 1):
        filled = sum(
            1
            for c in range(block["c1"], block["c2"] + 1)
            if ws.cell(row=r, column=c).value is not None
        )
        if filled >= 2:
            start = r
            break
    return (
        f"{get_column_letter(block['c1'])}{start}:"
        f"{get_column_letter(block['c2'])}{max_row}"
    )


def _explicit_range(target, texts: list[str]):
    """질문에 명시된 범위를 찾아 (시트, 범위)로 돌려준다. 없으면 None.

    시트명이 붙어 있으면 그 시트가 실제로 있을 때만 쓰고, 시트명이 없으면
    보이는 시트가 하나뿐일 때만 그 시트로 확정한다 — 여러 시트 중 어느
    시트인지 추정하지 않는다 (자동 블록 선택으로 폴백).
    """
    wb = _load(target, data_only=True)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    for text in texts:
        for match in _RANGE_RE.finditer(text):
            cell_range = match.group("range").replace("$", "").upper()
            sheet = match.group("sheet")
            if sheet and sheet in visible:
                return sheet, cell_range
            if not sheet and len(visible) == 1:
                return visible[0], cell_range
    return None


def _pick_table_block(target, texts: list[str]):
    """가장 큰 값-블록(헤더+데이터 2행 이상)을 고른다. 시트명 언급 시 그 시트로 한정."""
    wb = _load(target, data_only=True)
    sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    mentioned = [
        ws for ws in sheets if any(ws.title in text for text in texts)
    ]
    best = None
    for ws in mentioned or sheets:
        for block in _detect_blocks(ws):
            if block["rows"] < 2:
                continue
            score = block["rows"] * block["cols"]
            if best is None or score > best[0]:
                best = (score, ws.title, _trim_title_rows(ws, block))
    return (best[1], best[2]) if best else None


class AnalystNodes:
    def __init__(self, model) -> None:
        self.model = model

    def triage(self, state: AnalystState) -> dict[str, Any]:
        """SQL 분석 요청인지 일반 대화인지 분기한다.

        LLM 구조화 출력 1회로 판단하고, 실패하면 휴리스틱(파일 언급이
        있으면 analysis, 없으면 chat)으로 폴백한다.
        """
        _emit("triaging", "요청 유형을 판단하는 중")
        texts = _human_texts_newest_first(state)
        question = texts[0].strip() if texts else ""
        if not question:
            return {"question": question, "mode": "chat"}
        has_file = _find_target_file(texts) is not None
        context = _conversation_context(state)
        context_part = f"이전 대화:\n{context}\n" if context else ""
        try:
            decider = self.model.with_structured_output(TriageDecision)
            result = decider.invoke(
                [
                    SystemMessage(
                        content=(
                            "당신은 데이터 분석 에이전트의 라우터입니다. 사용자 "
                            "메시지가 표 데이터에 대한 집계·필터·정렬·계산 요청이면 "
                            "analysis, 인사·사용법 질문·기능 문의·이전 결과에 대한 "
                            "후속 설명처럼 SQL 실행이 필요 없는 대화면 chat으로 "
                            "분류하세요. 이전 대화의 분석을 새 조건으로 이어가는 "
                            "후속 요청(예: '그중 상위 3개만')은 analysis입니다."
                        )
                    ),
                    HumanMessage(
                        content=(
                            f"{context_part}"
                            f"메시지: {question}\n"
                            f"메시지에 분석 대상 파일 언급 존재: {has_file}"
                        )
                    ),
                ]
            )
            decision = (
                result
                if isinstance(result, TriageDecision)
                else TriageDecision.model_validate(result)
            )
            mode = decision.mode
        except Exception:
            mode = "analysis" if has_file else "chat"
        # 이전 턴이 남긴 error가 이번 턴 라우팅을 오염시키지 않게 리셋
        return {"question": question, "mode": mode, "error": None}

    def route_triage(self, state: AnalystState) -> Literal["analysis", "chat"]:
        return "chat" if state.get("mode") == "chat" else "analysis"

    def chat(self, state: AnalystState) -> dict[str, Any]:
        """SQL 없이 답하는 일반 대화 — 역할·사용법·파일 목록을 컨텍스트로 준다."""
        _emit("chatting", "일반 대화로 응답하는 중")
        response = self.model.invoke(
            [
                SystemMessage(
                    content=(
                        "당신은 '대형 엑셀 분석 Agent'입니다 — Excel·CSV 표를 "
                        "격리된 읽기 전용 DuckDB SQL로 집계·분석하는 데이터 분석 "
                        "모드입니다. 지금 요청은 SQL 실행이 필요 없는 일반 대화로 "
                        "분류되었습니다. 한국어로 간결히 답하고, 분석 방법을 물으면 "
                        "사용법을 안내하세요: 파일을 첨부하거나 파일명을 언급해 "
                        "질문하고, 표 위치가 애매하면 '시트명!A1:C50' 형식으로 "
                        "범위를 지정할 수 있습니다. 이전 분석 결과에 대한 후속 "
                        "질문이면 대화 맥락의 결과만으로 답하고, 새 계산이 필요하면 "
                        "분석 요청으로 다시 물어봐 달라고 안내하세요.\n\n"
                        f"[현재 볼 수 있는 파일]\n{list_workpapers.func()}"
                    )
                ),
                *state["messages"],
            ]
        )
        _emit("complete", "응답 완료")
        return {"messages": [response], "error": None}

    def inspect_data(self, state: AnalystState) -> dict[str, Any]:
        _emit("profiling", "대상 파일과 표 범위를 찾는 중")
        texts = _human_texts_newest_first(state)
        question = texts[0].strip() if texts else ""
        if not question:
            return {"error": "질문이 비어 있습니다.", "question": ""}

        target = _find_target_file(texts)
        if target is None:
            return {
                "question": question,
                "error": missing_file_message("분석할 Excel 파일을 찾지 못했습니다."),
            }

        picked = _explicit_range(target, texts)
        if picked:
            _emit("profiling", f"질문에 명시된 범위 사용: [{picked[0]}] {picked[1]}")
        else:
            picked = _pick_table_block(target, texts)
        if picked is None:
            wb = _load(target, data_only=True)
            sheets = ", ".join(
                ws.title for ws in wb.worksheets if ws.sheet_state == "visible"
            )
            return {
                "question": question,
                "error": (
                    f"'{target.name}'에서 표 형태의 데이터 블록(2행 이상)을 찾지 "
                    f"못했습니다. (시트: {sheets}) 분석할 범위를 '시트명!A1:C50' "
                    f"형식으로 지정해 다시 요청해 주세요."
                ),
            }
        sheet, cell_range = picked
        _emit("profiling", f"표 등록: {target.name} [{sheet}] {cell_range}")

        try:
            store = _store_cached(
                str(target), target.stat().st_mtime, sheet, cell_range, True
            )
        except ValueError as exc:
            return {"question": question, "error": str(exc)}

        described = store.connection.execute(f"DESCRIBE {TABLE}").fetchall()
        schema = "\n".join(f'- "{name}": {kind}' for name, kind, *_ in described)
        nulls = store.frame.isna().sum()
        sample = store.frame.head(5).astype(object)
        profile = {
            "row_count": len(store.frame),
            "column_count": len(store.frame.columns),
            "null_counts": {k: int(v) for k, v in nulls.items()},
            "original_headers": dict(store.column_map),
            "sample_rows": sample.where(sample.notna(), None).to_dict(orient="records"),
        }
        return {
            "question": question,
            "dataset": {"path": target.name, "sheet": sheet, "cell_range": cell_range},
            "schema": schema,
            "profile": profile,
            "attempts": 0,
            "error": None,
        }

    def route_inspect(self, state: AnalystState) -> Literal["plan", "fail"]:
        return "fail" if state.get("error") else "plan"

    def plan_sql(self, state: AnalystState) -> dict[str, Any]:
        _emit("planning", "질문을 읽기 전용 DuckDB SQL로 변환하는 중")
        return self._generate_sql(state, "Create the initial query.")

    def revise_sql(self, state: AnalystState) -> dict[str, Any]:
        attempt = state.get("attempts", 0) + 1
        _emit("revising_sql", "SQL 오류를 수정하는 중", attempt=attempt)
        update = self._generate_sql(
            state,
            f"Repair the previous SQL. Previous SQL: {state.get('sql', '')}\n"
            f"Error: {state.get('error', '')}",
        )
        update["attempts"] = attempt
        return update

    def _generate_sql(self, state: AnalystState, instruction: str) -> dict[str, Any]:
        planner = self.model.with_structured_output(SQLPlan)
        context = _conversation_context(state)
        context_part = (
            f"Recent conversation (for follow-up questions):\n{context}\n"
            if context
            else ""
        )
        result = planner.invoke(
            [
                SystemMessage(
                    content=(
                        f"You are a DuckDB analyst. Return exactly one read-only SELECT "
                        f"or WITH query against the {TABLE} table using only the listed "
                        f"columns. Column names may be Korean — always wrap them in "
                        f'double quotes (e.g. SELECT "금액_원" FROM {TABLE}). Never read '
                        f"files, URLs, extensions, catalogs, secrets, or environment "
                        f"data. Do not use markdown fences. The question may be a "
                        f"follow-up — resolve references like 'those', 'the top one' "
                        f"from the recent conversation."
                    )
                ),
                HumanMessage(
                    content=(
                        f"{context_part}"
                        f"Question: {state['question']}\nSchema:\n{state['schema']}\n"
                        f"Profile: {state['profile']!r}\nInstruction: {instruction}"
                    )
                ),
            ]
        )
        plan = result if isinstance(result, SQLPlan) else SQLPlan.model_validate(result)
        return {"sql": plan.sql, "error": None}

    def validate_sql(self, state: AnalystState) -> dict[str, Any]:
        _emit("validating_sql", "SQL 안전성과 테이블 범위를 검증하는 중")
        try:
            return {"sql": _validate_sql(state["sql"]), "error": None}
        except ValueError as exc:
            return {"error": str(exc)}

    def execute_sql(self, state: AnalystState) -> dict[str, Any]:
        _emit("executing", "격리된 DuckDB에서 쿼리를 실행하는 중")
        ds = state["dataset"]
        target = _base_dir() / ds["path"]
        try:
            store = _store_cached(
                str(target), target.stat().st_mtime, ds["sheet"], ds["cell_range"], True
            )
            # N+1 조회로 절단 여부를 감지한다 — answer 노드가 절단 사실을 밝혀야 하므로
            limited = f"SELECT * FROM ({state['sql']}) AS q LIMIT {MAX_RESULT_ROWS + 1}"
            cursor = store.connection.execute(limited)
            columns = [item[0] for item in cursor.description]
            rows = [list(row) for row in cursor.fetchall()]
            return {
                "result_columns": columns,
                "result_rows": rows[:MAX_RESULT_ROWS],
                "truncated": len(rows) > MAX_RESULT_ROWS,
                "error": None,
            }
        except Exception as exc:
            return {"error": f"SQL 실행 실패 — {exc}"}

    def route_sql(self, state: AnalystState) -> Literal["execute", "revise", "fail"]:
        if not state.get("error"):
            return "execute"
        if state.get("attempts", 0) < MAX_SQL_REVISIONS:
            return "revise"
        return "fail"

    def route_execution(self, state: AnalystState) -> Literal["answer", "revise", "fail"]:
        if not state.get("error"):
            return "answer"
        if state.get("attempts", 0) < MAX_SQL_REVISIONS:
            return "revise"
        return "fail"

    def answer(self, state: AnalystState) -> dict[str, Any]:
        _emit("answering", "실행 결과를 해석하는 중", rows=len(state["result_rows"]))
        ds = state["dataset"]
        table_md = _markdown(state["result_columns"], state["result_rows"])
        context = _conversation_context(state)
        context_part = f"이전 대화:\n{context}\n\n" if context else ""
        response = self.model.invoke(
            [
                SystemMessage(
                    content=(
                        "당신은 신입 회계사를 돕는 데이터 분석가입니다. 제공된 쿼리 "
                        "결과만으로 한국어로 답하세요. 후속 질문이면 이전 대화 맥락을 "
                        "이어서 답합니다. 결과가 비었거나 절단됐으면 "
                        "그 사실을 밝히고, 상관관계에서 인과를 추론하지 마세요. "
                        "재현을 위해 사용한 SQL을 ```sql 블록으로, 근거로 대상 "
                        "파일·시트·범위를 답변에 포함하세요."
                    )
                ),
                HumanMessage(
                    content=(
                        f"{context_part}"
                        f"질문: {state['question']}\n"
                        f"대상: {ds['path']} [{ds['sheet']}] {ds['cell_range']}\n"
                        f"SQL: {state['sql']}\n결과:\n{table_md}\n"
                        f"절단됨(상한 {MAX_RESULT_ROWS}행): {state.get('truncated', False)}\n"
                        f"프로필: {state['profile']!r}"
                    )
                ),
            ]
        )
        _emit("complete", "데이터 분석 완료")
        return {"messages": [response], "error": None}

    def fail(self, state: AnalystState) -> dict[str, Any]:
        message = state.get("error") or "데이터 분석에 실패했습니다."
        _emit("failed", message)
        return {"messages": [AIMessage(content=f"오류: {message}")]}


async def analyst(config: RunnableConfig):
    """요청 config로 모델을 정해 분석 그래프를 조립한다 (langgraph 서버가 호출)."""
    model_spec = (config.get("configurable") or {}).get("model", DEFAULT_MODEL)
    nodes = AnalystNodes(resolve_model(model_spec))

    builder = StateGraph(AnalystState)
    builder.add_node("triage", nodes.triage)
    builder.add_node("chat", nodes.chat)
    builder.add_node("inspect_data", nodes.inspect_data)
    builder.add_node("plan_sql", nodes.plan_sql)
    builder.add_node("validate_sql", nodes.validate_sql)
    builder.add_node("revise_sql", nodes.revise_sql)
    builder.add_node("execute_sql", nodes.execute_sql)
    builder.add_node("answer", nodes.answer)
    builder.add_node("fail", nodes.fail)

    builder.add_edge(START, "triage")
    builder.add_conditional_edges(
        "triage", nodes.route_triage, {"analysis": "inspect_data", "chat": "chat"}
    )
    builder.add_edge("chat", END)
    builder.add_conditional_edges(
        "inspect_data", nodes.route_inspect, {"plan": "plan_sql", "fail": "fail"}
    )
    builder.add_edge("plan_sql", "validate_sql")
    builder.add_conditional_edges(
        "validate_sql",
        nodes.route_sql,
        {"execute": "execute_sql", "revise": "revise_sql", "fail": "fail"},
    )
    builder.add_edge("revise_sql", "validate_sql")
    builder.add_conditional_edges(
        "execute_sql",
        nodes.route_execution,
        {"answer": "answer", "revise": "revise_sql", "fail": "fail"},
    )
    builder.add_edge("answer", END)
    builder.add_edge("fail", END)
    return builder.compile()
