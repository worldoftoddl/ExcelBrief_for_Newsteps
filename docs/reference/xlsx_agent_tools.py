# -*- coding: utf-8 -*-
"""
xlsx_agent_tools.py — 엑셀 '입체 이해' 에이전트용 도구 세트 (MVP 8종)

[ExcelBrief 저장소에서의 역할 — 이식 시 필독]
  이 파일은 드롭인 모듈이 아니라 **알고리즘 이식원**이다 (system_design 4.1).
  아래 설계 원칙 중 1(JSON 봉투)·3(세션+클로저)은 이 파일의 출신 맥락
  (단일 사용자 로컬 에이전트 세션)의 결정으로, ExcelBrief 방침과 다르다:

설계 원칙 (이 파일 자체의 맥락)
  1. 통일 봉투(envelope): 모든 도구가 같은 JSON 구조를 반환한다.
  2. 셀 원자(cell atom): 셀은 항상 {"addr","value","formula?"} 형태로만 표현한다.
  3. 세션 + 클로저 팩토리: 상태(2회 로드된 워크북, df 레지스트리)는
     WorkbookSession이 들고, LangChain 도구는 make_langchain_tools(session)이
     클로저로 캡처해 무상태 함수처럼 노출한다.
  4. 읽기 전용 / 절단 고지 / 주소 인용 강제.

의존성: openpyxl, pandas  (LangChain 연동 시 langchain-core)
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# ---------------------------------------------------------------------------
# 0. 공통 봉투 — 스키마의 심장
# ---------------------------------------------------------------------------

def envelope(
    data: Any = None,
    *,
    source: Optional[str] = None,
    truncated: Optional[dict] = None,
    note: Optional[str] = None,
    error: Optional[str] = None,
) -> str:
    """모든 도구의 유일한 반환 경로.

    ok        : 성공 여부. 예외는 절대 밖으로 던지지 않고 error 문자열로 승격한다
                (도구 예외가 ReAct 루프를 죽이면 안 되고, 에러도 Observation으로
                들어가야 에이전트가 자가수정한다).
    data      : 도구별 페이로드. 셀은 반드시 cell atom 형태.
    meta.source    : 이 데이터의 출처 주소(시트!범위). 최종 답변의 셀 인용 검증이
                     이 필드에서 나온다.
    meta.truncated : {"shown": n, "total": N, "hint": ...} — 조향 신호.
    meta.note      : 도구가 에이전트에게 주는 힌트(단위, 후속 도구 제안 등).
    """
    return json.dumps(
        {
            "ok": error is None,
            "data": data,
            "meta": {"source": source, "truncated": truncated, "note": note},
            "error": error,
        },
        ensure_ascii=False,
        default=str,
    )


def _guard(fn):
    """도구 본체 예외 → envelope(error=...) 변환 데코레이터."""

    def wrapper(*args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except Exception as e:  # noqa: BLE001 — 의도적 광역 포획
            return envelope(error=f"{type(e).__name__}: {e}")

    return wrapper


# ---------------------------------------------------------------------------
# 1. 세션 — 상태의 유일한 거처
# ---------------------------------------------------------------------------

@dataclass
class WorkbookSession:
    """워크북 1개에 대한 에이전트 세션.

    openpyxl은 수식과 캐시값을 한 번에 못 주므로 2회 로드해 둘 다 보관한다.
    frames에는 read_table이 등록한 DataFrame이 쌓이고 query가 이를 소비한다.
    """

    path: str
    wb_f: Any = field(default=None, repr=False)  # data_only=False → 수식 문자열
    wb_v: Any = field(default=None, repr=False)  # data_only=True  → 캐시된 값
    frames: dict[str, pd.DataFrame] = field(default_factory=dict)

    def open(self) -> "WorkbookSession":
        self.wb_f = load_workbook(self.path, data_only=False)
        self.wb_v = load_workbook(self.path, data_only=True)
        return self

    # -- 내부 헬퍼 --------------------------------------------------------
    def atom(self, sheet: str, coord: str) -> dict:
        """셀 원자: 모든 도구가 셀을 이 형태로만 내보낸다."""
        cf = self.wb_f[sheet][coord]
        cv = self.wb_v[sheet][coord]
        d: dict[str, Any] = {"addr": f"{sheet}!{coord}", "value": cv.value}
        if isinstance(cf.value, str) and cf.value.startswith("="):
            d["formula"] = cf.value
        return d


# ---------------------------------------------------------------------------
# 2. 내부 유틸
# ---------------------------------------------------------------------------

_REF_RE = re.compile(r"(?<![A-Z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)(?!\()")


def _to_r1c1(formula: str, base_row: int, base_col: int) -> str:
    """A1 → R1C1 정규화(상대참조만 오프셋화, $절대참조는 고정 좌표 유지).

    한계: 문자열 리터럴 안의 유사 참조("A1 텍스트")까지 치환될 수 있다.
    구조 파악용 압축이 목적이므로 허용.
    """

    def repl(m: re.Match) -> str:
        dc, col, dr, row = m.groups()
        ci, ri = column_index_from_string(col), int(row)
        c = f"C{ci}" if dc else ("C" if ci == base_col else f"C[{ci - base_col}]")
        r = f"R{ri}" if dr else ("R" if ri == base_row else f"R[{ri - base_row}]")
        return r + c

    return _REF_RE.sub(repl, formula)


def _detect_tables(ws) -> list[dict]:
    """빈 행 2개 이상을 경계로 값-블록(=테이블 후보)을 분할."""
    row_span: dict[int, tuple[int, int]] = {}
    for row in ws.iter_rows():
        cells = [c for c in row if c.value is not None]
        if cells:
            row_span[cells[0].row] = (
                min(c.column for c in cells),
                max(c.column for c in cells),
            )
    tables, cur = [], []
    prev = None
    for r in sorted(row_span):
        if prev is not None and r - prev > 2:  # 공백 2행 이상 → 새 블록
            tables.append(cur)
            cur = []
        cur.append(r)
        prev = r
    if cur:
        tables.append(cur)

    out = []
    for rows in tables:
        c1 = min(row_span[r][0] for r in rows)
        c2 = max(row_span[r][1] for r in rows)
        ref = f"{get_column_letter(c1)}{rows[0]}:{get_column_letter(c2)}{rows[-1]}"
        out.append({"ref": ref, "rows": len(rows), "cols": c2 - c1 + 1})
    return out


def _iter_ref(ws, ref: Optional[str]):
    """ref('B2' | 'A1:D9' | None=전체)를 셀 이터레이터로 정규화."""
    if ref is None:
        for row in ws.iter_rows():
            yield from row
        return
    got = ws[ref]
    if hasattr(got, "coordinate"):  # 단일 셀
        yield got
    else:
        for row in got:
            yield from row


# ---------------------------------------------------------------------------
# 3. 도구 8종 (세션을 첫 인자로 받는 순수 함수)
# ---------------------------------------------------------------------------

@_guard
def workbook_overview(s: WorkbookSession) -> str:
    """[반드시 첫 호출] 시트 목록·크기·수식 밀도 → 데이터 흐름 지도."""
    sheets = []
    for ws in s.wb_f.worksheets:
        n_val = n_frm = n_x = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_frm += 1
                    if "!" in c.value:
                        n_x += 1
                else:
                    n_val += 1
        sheets.append(
            {
                "sheet": ws.title,
                "dim": ws.dimensions,
                "state": ws.sheet_state,  # visible/hidden/veryHidden
                "value_cells": n_val,
                "formula_cells": n_frm,
                "cross_sheet_refs": n_x,
                "merged_ranges": len(ws.merged_cells.ranges),
            }
        )
    note = (
        "cross_sheet_refs가 높고 value_cells가 낮은 시트=출력물, "
        "value_cells 위주=원천 데이터일 가능성. 다음: sheet_overview로 드릴다운."
    )
    return envelope(sheets, source=s.path.rsplit("/", 1)[-1], note=note)


@_guard
def sheet_overview(s: WorkbookSession, sheet: str) -> str:
    """시트 1장의 지도: 테이블 경계 + 헤더 미리보기 + 병합/숨김."""
    ws_f, ws_v = s.wb_f[sheet], s.wb_v[sheet]
    tables = _detect_tables(ws_f)
    for t in tables:
        first_row = int(re.search(r"(\d+)", t["ref"]).group(1))
        preview = []
        for c in _iter_ref(ws_v, t["ref"].split(":")[0] + ":" +
                           get_column_letter(min(column_index_from_string(
                               re.match(r"([A-Z]+)", t["ref"].split(":")[1]).group(1)),
                               column_index_from_string(
                               re.match(r"([A-Z]+)", t["ref"]).group(1)) + 9))
                           + str(first_row)):
            if c.value is not None:
                preview.append(str(c.value)[:20])
        t["header_preview"] = preview
    hidden_rows = [r for r, d in ws_f.row_dimensions.items() if d.hidden]
    hidden_cols = [c for c, d in ws_f.column_dimensions.items() if d.hidden]
    data = {
        "tables": tables,
        "merged": [str(r) for r in ws_f.merged_cells.ranges][:20],
        "hidden_rows": hidden_rows[:20],
        "hidden_cols": hidden_cols[:20],
    }
    return envelope(
        data,
        source=f"{sheet}!{ws_f.dimensions}",
        note="테이블 ref를 read_table 또는 get_range에 그대로 넘길 것.",
    )


def _color_tag(color) -> Optional[str]:
    """openpyxl 색 객체 → 짧은 태그. rgb/theme/indexed 3계열 분기.

    theme 색은 `.rgb`가 문자열이 아니라 접근 시 오류 객체를 반환하므로
    반드시 `type`으로 먼저 분기한다 (실측: 감사조서 표준 서식의 음영은
    전부 theme 계열). theme는 T{번호}±{tint}, indexed는 I{번호} 표기.
    """
    if color is None:
        return None
    if color.type == "rgb":
        rgb = color.rgb
        return rgb if isinstance(rgb, str) and rgb != "00000000" else None
    if color.type == "theme":
        tint = round(color.tint or 0, 2)
        return f"T{color.theme}{tint:+g}" if tint else f"T{color.theme}"
    if color.type == "indexed" and color.indexed not in (64, 65):  # 64/65=자동
        return f"I{color.indexed}"
    return None


@_guard
def get_range(
    s: WorkbookSession, sheet: str, ref: str, with_format: bool = False, max_cells: int = 200
) -> str:
    """범위의 셀 원자 그리드. 사람의 '스크롤'에 대응."""
    ws_f = s.wb_f[sheet]
    cells, total = [], 0
    for c in _iter_ref(ws_f, ref):
        total += 1
        if c.value is None or len(cells) >= max_cells:
            continue
        a = s.atom(sheet, c.coordinate)
        if with_format:
            fmt = []
            if c.font is not None and c.font.bold:
                fmt.append("bold")
            fill_tag = (
                _color_tag(c.fill.fgColor)
                if c.fill is not None and c.fill.patternType
                else None
            )
            if fill_tag:
                fmt.append(f"fill:{fill_tag}")
            font_tag = _color_tag(c.font.color) if c.font is not None else None
            # 기본 글자색(검정 FF000000, 테마 0/1 무틴트)은 잡음이므로 생략
            if font_tag and font_tag not in ("FF000000", "T0", "T1"):
                fmt.append(f"color:{font_tag}")
            if fmt:
                a["format"] = fmt
        cells.append(a)
    trunc = None
    if total > max_cells:
        trunc = {"shown": len(cells), "total": total, "hint": "범위를 좁혀 재호출"}
    return envelope(cells, source=f"{sheet}!{ref}", truncated=trunc)


@_guard
def find(
    s: WorkbookSession, pattern: str, sheet: Optional[str] = None, regex: bool = False, max_hits: int = 50
) -> str:
    """값·수식 전체에서 Ctrl+F. 히트 좌표만 반환한다."""
    rx = re.compile(pattern if regex else re.escape(pattern), re.IGNORECASE)
    hits = []
    sheets = [sheet] if sheet else s.wb_f.sheetnames
    for sh in sheets:
        for row in s.wb_f[sh].iter_rows():
            for c in row:
                if len(hits) >= max_hits:
                    break
                if c.value is not None and rx.search(str(c.value)):
                    hits.append(s.atom(sh, c.coordinate))
            if len(hits) >= max_hits:
                break
        if len(hits) >= max_hits:
            break
    trunc = {"shown": len(hits), "total": "≥shown", "hint": "패턴을 좁힐 것"} if len(hits) >= max_hits else None
    return envelope(hits, source=f"search:{pattern}", truncated=trunc)


@_guard
def get_formulas(s: WorkbookSession, sheet: str, ref: Optional[str] = None, max_cells: int = 100) -> str:
    """수식 셀만 골라 값+수식 병기로 반환. '수식 입력줄'에 대응."""
    out, total = [], 0
    for c in _iter_ref(s.wb_f[sheet], ref):
        if isinstance(c.value, str) and c.value.startswith("="):
            total += 1
            if len(out) < max_cells:
                out.append(s.atom(sheet, c.coordinate))
    trunc = {"shown": len(out), "total": total, "hint": "ref로 좁힐 것"} if total > max_cells else None
    return envelope(out, source=f"{sheet}!{ref or 'ALL'}", truncated=trunc)


@_guard
def formula_map(s: WorkbookSession, sheet: str, top: int = 20) -> str:
    """R1C1 정규화 패턴 압축 + 패턴 이탈(수식 지대 내 하드코딩 숫자) 검출.

    이탈 판정 휴리스틱: '같은 행과 같은 열 모두에 수식이 존재'하는 자리의
    숫자 상수. 오류일 수도, 설계된 입력 셀일 수도 있으므로 판정은
    에이전트(도메인 지식)의 몫 — 도구는 검출만 한다.
    """
    ws = s.wb_f[sheet]
    patterns: dict[str, list[str]] = defaultdict(list)
    frows, fcols = set(), set()
    numerics: list = []
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                patterns[_to_r1c1(c.value, c.row, c.column)].append(c.coordinate)
                frows.add(c.row)
                fcols.add(c.column)
            elif isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                numerics.append(c)
    deviants = [
        s.atom(sheet, c.coordinate) for c in numerics if c.row in frows and c.column in fcols
    ][:30]
    ranked = sorted(patterns.items(), key=lambda kv: -len(kv[1]))
    data = {
        "n_formulas": sum(len(v) for v in patterns.values()),
        "n_patterns": len(patterns),
        "top_patterns": [
            {"r1c1": k[:100], "count": len(v), "examples": v[:3]} for k, v in ranked[:top]
        ],
        "deviants": deviants,
    }
    return envelope(
        data,
        source=f"{sheet}!ALL",
        note="deviants는 오류/입력셀 양쪽 가능 — 도메인 판단 필요.",
    )


@_guard
def read_table(
    s: WorkbookSession, sheet: str, ref: str, header_rows: int = 1, name: Optional[str] = None
) -> str:
    """좌표 → DataFrame 파싱·등록. 셀 세계에서 테이블 세계로의 인계."""
    ws = s.wb_v[sheet]
    grid = [[c.value for c in row] for row in ws[ref]]
    if len(grid) <= header_rows:
        return envelope(error="header_rows가 데이터보다 큼")
    headers = grid[:header_rows]
    # 병합으로 비는 헤더는 왼쪽 값으로 전진 채움 후 다층이면 '_' 결합
    for h in headers:
        for i in range(1, len(h)):
            if h[i] is None:
                h[i] = h[i - 1]
    cols = [
        "_".join(str(h[i]) for h in headers if h[i] is not None) or f"col{i}"
        for i in range(len(headers[0]))
    ]
    # 병합 헤더 전진 채움의 부작용: 열명 중복 → 위치 접근이 깨지므로 디듀프
    seen: dict[str, int] = {}
    for i, cname in enumerate(cols):
        if cname in seen:
            seen[cname] += 1
            cols[i] = f"{cname}.{seen[cname]}"
        else:
            seen[cname] = 0
    df = pd.DataFrame(grid[header_rows:], columns=cols)
    # "1,250,000" 문자열 → 숫자 강제 (70% 이상 변환되면 채택). 위치 기반 접근.
    for i in range(df.shape[1]):
        col = df.iloc[:, i]
        if col.dtype == object:
            conv = pd.to_numeric(
                col.astype(str).str.replace(",", "", regex=False), errors="coerce"
            )
            if conv.notna().sum() >= max(1, int(0.7 * max(col.notna().sum(), 1))):
                df.iloc[:, i] = conv
    key = name or f"{sheet}_{ref.replace(':', '_')}"
    s.frames[key] = df
    data = {
        "frame": key,
        "shape": list(df.shape),
        "columns": cols,
        "head": df.head(3).to_dict(orient="records"),
    }
    return envelope(data, source=f"{sheet}!{ref}", note=f"query에서 frames['{key}']로 사용 가능.")


@_guard
def query(s: WorkbookSession, expr: str, max_chars: int = 2000) -> str:
    """등록된 DataFrame에 대한 pandas 표현식 평가. 계산 위임의 종착지.

    보안 주의: eval 기반 경량 샌드박스(로컬 데모용). 외부 노출 서비스라면
    subprocess 격리나 RestrictedPython으로 교체할 것.
    """
    if not s.frames:
        return envelope(error="등록된 frame 없음 — read_table 선행 필요")
    ns = {"pd": pd, "frames": s.frames, **s.frames}
    result = eval(expr, {"__builtins__": {}}, ns)  # noqa: S307
    text = result.to_string() if isinstance(result, (pd.DataFrame, pd.Series)) else repr(result)
    trunc = None
    if len(text) > max_chars:
        text, trunc = text[:max_chars], {"shown": max_chars, "total": len(text), "hint": "집계로 좁힐 것"}
    return envelope(
        {"result": text}, source=f"query:{expr[:60]}", truncated=trunc,
        note=f"사용 가능 frame: {list(s.frames)}",
    )


# ---------------------------------------------------------------------------
# 4. LangChain 도구 팩토리 — deepagents/LangGraph 배선 지점
# ---------------------------------------------------------------------------

def make_langchain_tools(session: WorkbookSession) -> list:
    """세션을 클로저로 캡처한 @tool 리스트. create_deep_agent(tools=...)에 그대로."""
    from langchain_core.tools import tool

    @tool
    def xlsx_overview() -> str:
        """[필수 첫 호출] 워크북 전체 지도: 시트별 크기·값/수식 밀도·시트간 참조 수.
        데이터 흐름(원천→가공→출력) 추정의 출발점."""
        return workbook_overview(session)

    @tool
    def xlsx_sheet(sheet: str) -> str:
        """시트 1장의 지도: 테이블 경계·헤더 미리보기·병합·숨김 행렬.
        반환된 테이블 ref는 xlsx_range/xlsx_read_table에 그대로 넘긴다."""
        return sheet_overview(session, sheet)

    @tool
    def xlsx_range(sheet: str, ref: str, with_format: bool = False) -> str:
        """범위(예 'A1:D30')의 값을 셀 주소와 함께 읽는다. 사람의 스크롤에 대응.
        truncated가 오면 범위를 좁혀 재호출."""
        return get_range(session, sheet, ref, with_format)

    @tool
    def xlsx_find(pattern: str, sheet: str = "") -> str:
        """전 시트(또는 지정 시트)에서 텍스트/수식 검색. Ctrl+F에 대응.
        계정명·항목명으로 위치를 찾을 때 전체 덤프 대신 이것을 쓴다."""
        return find(session, pattern, sheet or None)

    @tool
    def xlsx_formulas(sheet: str, ref: str = "") -> str:
        """범위 내 수식 셀만 값+수식 병기로 반환. 수식 입력줄 클릭에 대응."""
        return get_formulas(session, sheet, ref or None)

    @tool
    def xlsx_formula_map(sheet: str) -> str:
        """시트 전체 수식을 R1C1 패턴으로 압축하고 패턴 이탈(수식 지대 내
        하드코딩 숫자)을 검출. 시트의 계산 골격 파악과 감사에 사용."""
        return formula_map(session, sheet)

    @tool
    def xlsx_read_table(sheet: str, ref: str, header_rows: int = 1, name: str = "") -> str:
        """테이블 범위를 DataFrame으로 파싱·등록. 이후 xlsx_query에서 집계 가능."""
        return read_table(session, sheet, ref, header_rows, name or None)

    @tool
    def xlsx_query(expr: str) -> str:
        """등록된 frame에 pandas 표현식 실행(예 "frames['CF_t1'].sum()").
        합계·증감·비율 등 모든 계산은 암산 대신 반드시 이 도구로."""
        return query(session, expr)

    return [
        xlsx_overview, xlsx_sheet, xlsx_range, xlsx_find,
        xlsx_formulas, xlsx_formula_map, xlsx_read_table, xlsx_query,
    ]


# 에이전트 시스템 프롬프트에 붙일 탐색 정책 조각
POLICY_PROMPT = """[엑셀 탐색 정책]
1. 파일을 열면 반드시 xlsx_overview → xlsx_sheet 순으로 지도를 확보한 뒤 드릴다운한다.
2. 모든 수치 주장에는 근거 셀 주소(시트!주소)를 인용한다. 인용할 수 없는 수치는 말하지 않는다.
3. 합계·증감·비율 등 계산은 암산하지 않는다 — xlsx_query 또는 원본 셀 재조회로 확인한다.
4. 응답의 meta.truncated를 보면 범위를 좁혀 재호출한다.
5. 발견한 구조(테이블 위치, 헤더 계층, 수식 패턴, 단위)는 스크래치패드 파일에 갱신하고,
   같은 범위를 재조회하기 전에 스크래치패드를 먼저 확인한다.
6. 최종 답변 전, 핵심 수치 1개 이상을 재조회로 검산한다."""


# ---------------------------------------------------------------------------
# 5. 데모
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    path = sys.argv[1]
    s = WorkbookSession(path).open()

    print("--- workbook_overview ---")
    print(workbook_overview(s)[:600], "\n")

    ov = json.loads(workbook_overview(s))
    target = max(ov["data"], key=lambda x: x["formula_cells"])["sheet"]

    print(f"--- sheet_overview({target}) ---")
    print(sheet_overview(s, target)[:600], "\n")

    print(f"--- formula_map({target}) ---")
    print(formula_map(s, target)[:800], "\n")

    print("--- find('감가상각비') ---")
    print(find(s, "감가상각비")[:500], "\n")

    tbl = json.loads(sheet_overview(s, target))["data"]["tables"]
    if tbl:
        ref = tbl[-1]["ref"]
        print(f"--- read_table({target}, {ref}) ---")
        print(read_table(s, target, ref, header_rows=1, name="t1")[:600], "\n")
        print("--- query ---")
        print(query(s, "frames['t1'].shape")[:300])