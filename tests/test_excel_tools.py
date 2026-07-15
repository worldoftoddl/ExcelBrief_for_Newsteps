"""Excel 탐색 도구 테스트 — 한공회 공식 조서 서식(실파일) 기준.

실파일에 없는 요소(메모·숨김·유효성)는 심은 파일(planted_dir)로 검증한다.
"""

import os
import re

import pytest

from agent.tools.excel import (
    excel_find,
    excel_formula_map,
    excel_get_annotations,
    excel_read_range,
    excel_sheet_stats,
    excel_workbook_overview,
    list_workpapers,
)

WP_계약 = "감사조서서식_1100~1300 감사계약.xlsx"
WP_3650 = "감사조서서식_3650 감사 전 재무제표 확인.xlsx"
WP_4000 = "감사조서서식_4000 계정별 실증절차 (KIFRS용) 2025.xlsx"
WP_대응 = "감사조서서식_3100-3800 위험에 대한 대응 2025.xlsx"


@pytest.fixture()
def planted_dir(tmp_path, monkeypatch):
    """실파일에 없는 요소(메모·숨김·유효성·정의된 이름·파이프 값)를 심은 워크북."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "본문"
    ws["A1"], ws["B1"] = "항목", "a|b"
    ws["B2"] = 100
    ws["B2"].comment = Comment("차이 조정명세 수령 예정", "인차지B")
    ws["C3"] = 200
    ws["C3"].comment = Comment("전기 대비 급증 — 소명 필요", "담당A")
    ws.column_dimensions["D"].hidden = True
    ws.row_dimensions[5].hidden = True
    dv = DataValidation(type="list", formula1='"완료,미완료"')
    ws.add_data_validation(dv)
    dv.add("E2:E4")
    prior = wb.create_sheet("전기자료")
    prior["A1"] = "prior"
    prior.sheet_state = "hidden"
    wb.defined_names["검토범위"] = DefinedName("검토범위", attr_text="본문!$A$1:$B$2")
    wb.save(tmp_path / "심은조서.xlsx")
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    return tmp_path


# ── list_workpapers ──────────────────────────────────────────
def test_list_workpapers_lists_official_forms():
    out = list_workpapers.invoke({})
    assert WP_계약 in out and WP_3650 in out and WP_4000 in out


# ── excel_workbook_overview ──────────────────────────────────
def test_overview_shows_sheets_and_merges():
    out = excel_workbook_overview.invoke({"path": WP_3650})
    assert "3650" in out and "3650A 신규" in out
    assert "49행 × 16열" in out
    assert "병합" in out


def test_overview_loads_phonetic_file():
    """openpyxl phonetic 속성 회귀 테스트 — 패치 없이는 로드가 죽는 파일."""
    out = excel_workbook_overview.invoke({"path": WP_4000})
    assert "시트 36개" in out
    assert "실증절차" in out


def test_overview_missing_file_suggests_candidates():
    out = excel_workbook_overview.invoke({"path": "없는파일.xlsx"})
    assert "오류" in out
    assert WP_3650 in out  # 후보 목록 제시


def test_overview_blocks_path_escape():
    out = excel_workbook_overview.invoke({"path": "../탈출.xlsx"})
    assert "오류" in out


# ── excel_read_range ─────────────────────────────────────────
def test_read_range_values_as_markdown():
    out = excel_read_range.invoke(
        {"path": WP_3650, "sheet": "3650", "cell_range": "A1:C3"}
    )
    assert "감사 전 재무제표 확인" in out
    assert "|" in out  # 마크다운 표


def test_read_range_formulas_mode_shows_cross_sheet_refs():
    out = excel_read_range.invoke(
        {"path": WP_계약, "sheet": "1200", "cell_range": "A4:D5", "mode": "formulas"}
    )
    assert "='1100'!B4" in out  # 시트 간 참조 수식


def test_read_range_cell_cap_returns_guidance():
    # 3650A 신규: 113행 × 13열 = 1,469셀 > 500
    out = excel_read_range.invoke(
        {"path": WP_3650, "sheet": "3650A 신규", "cell_range": "A1:M113"}
    )
    assert "500" in out and "나눠" in out  # 예외가 아니라 분할 안내


def test_read_range_bad_sheet_is_error_text():
    out = excel_read_range.invoke(
        {"path": WP_3650, "sheet": "없는시트", "cell_range": "A1:B2"}
    )
    assert "오류" in out


# ── excel_find ───────────────────────────────────────────────
def test_find_locates_cell_across_sheets():
    out = excel_find.invoke({"path": WP_3650, "query": "감사 전 재무제표"})
    assert "3650!A2" in out


def test_find_scoped_to_sheet():
    out = excel_find.invoke(
        {"path": WP_3650, "query": "감사 전 재무제표", "sheet": "3650"}
    )
    assert "3650!A2" in out
    assert "3650A 신규!" not in out


# ── excel_sheet_stats ────────────────────────────────────────
def test_sheet_stats_counts_formulas():
    out = excel_sheet_stats.invoke({"path": WP_계약, "sheet": "1200"})
    assert "수식 셀: 3" in out  # ='1100'!… 3건


# ── v2: 서식 채널 (mode="format") ────────────────────────────
def test_format_mode_marks_bold():
    out = excel_read_range.invoke(
        {"path": WP_3650, "sheet": "3650A 신규", "cell_range": "A2:A2", "mode": "format"}
    )
    assert "[B" in out  # 볼드 표기


def test_format_mode_marks_theme_fill():
    out = excel_read_range.invoke(
        {"path": WP_3650, "sheet": "3650A 신규", "cell_range": "A19:A19", "mode": "format"}
    )
    assert "F:T0" in out  # theme 계열 배경 — rgb만 처리하면 잡히지 않음


# ── v2: 의도 채널 (excel_get_annotations) ────────────────────
def test_annotations_full_recovery_on_planted(planted_dir):
    out = excel_get_annotations.invoke({"path": "심은조서.xlsx", "sheet": "본문"})
    assert "인차지B" in out and "차이 조정명세" in out  # 메모 회수
    assert "담당A" in out
    assert "D" in out and "5" in out  # 숨김 열·행
    assert "전기자료" in out  # 숨김 시트
    assert "E2:E4" in out  # 데이터 유효성
    assert "검토범위" in out  # 정의된 이름


def test_annotations_clean_real_file_says_none():
    out = excel_get_annotations.invoke({"path": WP_3650, "sheet": "3650"})
    assert "없음" in out


# ── v2: 수식 지도 (excel_formula_map) ────────────────────────
def test_formula_map_compresses_patterns():
    out = excel_formula_map.invoke({"path": WP_대응, "sheet": "3500"})
    assert "수식 4" in out and "패턴 2종" in out
    assert "B3" in out  # 패턴 예시 좌표


# ── v2: overview 블록 감지·수식 밀도 ─────────────────────────
def test_overview_reports_formula_density():
    out = excel_workbook_overview.invoke({"path": WP_계약})
    assert "수식 3" in out  # 1200/1300 시트


def test_overview_detects_blocks():
    out = excel_workbook_overview.invoke({"path": WP_3650})
    assert "블록" in out
    assert re.search(r"[A-Z]+28:[A-Z]+31", out)  # 3650 둘째 블록 (28~31행)


# ── v2: 잔손질 ───────────────────────────────────────────────
def test_find_formulas_mode_traces_cross_refs():
    out = excel_find.invoke({"path": WP_계약, "query": "'1100'!", "mode": "formulas"})
    assert "1200!B4" in out


def test_pipe_escaped_in_markdown(planted_dir):
    out = excel_read_range.invoke(
        {"path": "심은조서.xlsx", "sheet": "본문", "cell_range": "B1:B1"}
    )
    assert r"a\|b" in out


def test_workbook_cache_hits():
    from agent.tools.excel import _load_cached

    excel_workbook_overview.invoke({"path": WP_3650})
    before = _load_cached.cache_info().hits
    excel_workbook_overview.invoke({"path": WP_3650})
    assert _load_cached.cache_info().hits > before


# ── 에이전트 통합 (개요→정독 탐색 순서) ──────────────────────
@pytest.mark.skipif(not os.environ.get("ANTHROPIC_API_KEY"), reason="API 키 없음")
async def test_agent_explores_overview_first():
    from agent.graph import graph

    g = await graph({"configurable": {}})
    result = await g.ainvoke(
        {
            "messages": [
                {
                    "role": "user",
                    "content": f"{WP_3650} 파일이 어떤 조서인지 간단히 설명해줘.",
                }
            ]
        },
        config={"recursion_limit": 20},
    )
    tool_calls = [
        tc["name"]
        for m in result["messages"]
        for tc in (getattr(m, "tool_calls", None) or [])
    ]
    assert "excel_workbook_overview" in tool_calls
    assert result["messages"][-1].content
