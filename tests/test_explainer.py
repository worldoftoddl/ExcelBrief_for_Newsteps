"""explainer 그래프 테스트 — 분기·파일 탐지·렌더는 단위로,
전체 파이프라인은 실제 API(skipif)로 검증한다. 증거 수집·인용 확정은
공용 계층(evidence·standards_lookup) 테스트가 담당한다."""

import os

import pytest
from langchain_core.messages import HumanMessage
from openpyxl import Workbook

from agent.explainer import (
    ExplainerNodes,
    ProcedureNote,
    TermNote,
    WorkpaperBrief,
    _render_brief,
    explainer,
)

FILE = "표본_해설조서.xlsx"
SHEET = "5100"


@pytest.fixture()
def explain_dir(tmp_path, monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "현금 실사 조서"
    ws["A2"] = "작성자"
    ws["B2"] = "김신입"
    rows = [
        ["항목", "장부", "실사"],
        ["소액현금", 100, 100],
        ["보통예금", 900, 900],
    ]
    for i, row in enumerate(rows, start=5):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    ws["B8"] = "=SUM(B6:B7)"
    wb.save(tmp_path / FILE)
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    return tmp_path


def _sample_brief() -> WorkpaperBrief:
    return WorkpaperBrief(
        workpaper_purpose="현금 실사 결과를 장부와 대사하는 조서",
        sheet_roles=[f"{SHEET}: 실사 대사표"],
        performed_procedures=[
            ProcedureNote(
                procedure="현금 실사 대사",
                location=f"{SHEET} 시트의 실사 대사표(A5:C7)",
                interpretation="장부금액과 실사금액을 항목별로 비교한다",
                assertion="존재성",
                risk_addressed="실재하지 않는 현금이 장부에 남는다",
                standards_query="현금 실사 감사증거",
                source_hint="감사기준",
            )
        ],
        reading_tips=["B8 합계 수식이 대사표를 집계한다"],
        open_items=[],
        terms=[TermNote(term="실사", explanation="자산을 직접 확인하는 절차")],
        overall="작성 초기 단계의 실사 조서",
    )


def test_triage_fallback_heuristic(explain_dir):
    nodes = ExplainerNodes(model=None)
    update = nodes.triage(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}] 해설해줘")]}
    )
    assert update["mode"] == "explain" and update["error"] is None
    assert nodes.route_triage(update) == "explain"

    update = nodes.triage({"messages": [HumanMessage(content="안녕, 뭘 해줘?")]})
    assert update["mode"] == "chat"
    assert nodes.route_triage(update) == "chat"


def test_locate_without_file_lists_candidates(explain_dir):
    nodes = ExplainerNodes(model=None)
    update = nodes.locate({"messages": [HumanMessage(content="조서 해설해줘")]})
    assert update["error"] and FILE in update["error"]
    assert nodes.route_locate(update) == "fail"


def test_collect_populates_scope(explain_dir):
    nodes = ExplainerNodes(model=None)
    update = nodes.collect({"path": FILE})
    assert update["error"] is None
    assert "[서명란 스캔]" in update["evidence"]
    assert update["examined"] == [SHEET]
    assert nodes.route_collect(update) == "explain"


def test_render_brief_sections_and_citation(explain_dir):
    brief = _sample_brief()
    brief.performed_procedures[0].citation = "감사기준서 501 문단 A1"
    brief.performed_procedures[0].citation_cid = "KSA::501::A1"
    out = _render_brief(FILE, brief, [SHEET], [])
    for section in ("① 이 조서는 무엇인가", "② 시트 구성", "③ 수행된 절차 해설",
                    "④ 조서 읽는 법", "⑤ 미완·후속 확인 항목", "⑥ 용어 풀이",
                    "⑦ 한눈 요약", "점검 범위", "근거 목록"):
        assert section in out
    assert "근거: 감사기준서 501 문단 A1" in out
    assert "`KSA::501::A1`" in out
    assert "**실사**: 자산을 직접 확인하는 절차" in out
    # 하위 불릿 구조 — 마크다운에서 줄이 문단으로 접히지 않아야 한다
    assert "\n  - _주장: 존재성 · 대응 위험: 실재하지 않는 현금이 장부에 남는다_" in out
    assert "\n  - 장부금액과 실사금액을 항목별로 비교한다" in out
    assert "\n  - 근거: 감사기준서 501 문단 A1" in out


def test_brief_coerces_weak_model_strings():
    """약한 모델이 문자열 리스트를 내도 스키마로 승격된다."""
    brief = WorkpaperBrief.model_validate(
        {
            "workpaper_purpose": "p",
            "sheet_roles": [],
            "performed_procedures": ["현금 실사 대사 수행"],
            "reading_tips": [],
            "open_items": [],
            "terms": ["실사: 자산 직접 확인"],
            "overall": "o",
        }
    )
    assert brief.performed_procedures[0].procedure == "현금 실사 대사 수행"
    assert brief.performed_procedures[0].location == "(미표기)"
    assert brief.terms[0].term == "실사"


async def test_cite_attaches_citation(explain_dir, monkeypatch):
    from types import SimpleNamespace

    import agent.explainer as explainer_mod

    async def search_fn(**kwargs):
        return '{"results": [{"cid": "KSA::501::A1", "display": "감사기준서 501 문단 A1(적용자료)"}]}'

    async def para_fn(**kwargs):
        return '{"paragraphs": [{"cid": "KSA::501::A1", "display": "감사기준서 501 문단 A1(적용자료)"}]}'

    async def fake_tools():
        return [
            SimpleNamespace(name="standards_search", coroutine=search_fn),
            SimpleNamespace(name="standards_get_paragraph", coroutine=para_fn),
        ]

    monkeypatch.setattr(explainer_mod, "get_standards_tools", fake_tools)
    nodes = ExplainerNodes(model=None)
    update = await nodes.cite({"brief": _sample_brief().model_dump()})
    brief = WorkpaperBrief.model_validate(update["brief"])
    assert brief.performed_procedures[0].citation_cid == "KSA::501::A1"


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_full_pipeline_with_real_model(explain_dir):
    graph = await explainer(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}]\n이 조서 해설해줘")]}
    )
    text = result["messages"][-1].content
    assert "# 조서 해설" in text and "⑦ 한눈 요약" in text
    assert result["error"] is None
