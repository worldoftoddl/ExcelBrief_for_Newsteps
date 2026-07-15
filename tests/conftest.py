"""테스트 픽스처 — 조서 모사·범용 xlsx를 임시 폴더에 생성하고 WORKPAPERS_DIR로 지정."""

import os

import pytest
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()  # skipif(ANTHROPIC_API_KEY)가 수집 시점에 .env를 보도록 선행 로드


def _make_workpaper(path):
    """감사조서 모사: 병합 셀·수식·틱마크·대용량 시트 포함."""
    wb = Workbook()

    lead = wb.active
    lead.title = "Lead"
    lead["A1"] = "매출채권 Lead Schedule"
    lead.merge_cells("A1:C1")
    lead["A3"], lead["B3"], lead["C3"] = "계정과목", "전기말", "당기말"
    lead["A4"], lead["B4"], lead["C4"] = "매출채권", 1000, 1500
    lead["A5"], lead["B5"], lead["C5"] = "대손충당금", -30, -45
    lead["B6"], lead["C6"] = "=SUM(B4:B5)", "=SUM(C4:C5)"

    detail = wb.create_sheet("명세")
    detail["A1"], detail["B1"], detail["C1"] = "거래처", "금액", "틱마크"
    detail["A2"], detail["B2"], detail["C2"] = "한빛상사", 420, "✓"
    detail["A3"], detail["B3"] = "동서물산", 250

    big = wb.create_sheet("빅시트")
    for row in range(1, 31):          # 30행 × 20열 = 600셀 (500셀 상한 초과용)
        for col in range(1, 21):
            big.cell(row=row, column=col, value=row * col)

    wb.save(path)


def _make_generic(path):
    """범용(비감사) Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "판매"
    ws.append(["product", "qty", "price"])
    ws.append(["노트북", 3, 1_200_000])
    ws.append(["모니터", 5, 350_000])
    wb.save(path)


@pytest.fixture(scope="session", autouse=True)
def workpapers_dir(tmp_path_factory):
    d = tmp_path_factory.mktemp("workpapers")
    _make_workpaper(d / "조서_테스트.xlsx")
    _make_generic(d / "범용_판매데이터.xlsx")
    os.environ["WORKPAPERS_DIR"] = str(d)
    return d
