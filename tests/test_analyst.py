"""analyst 그래프 테스트 — 파일 탐지·블록 선택·비LLM 노드는 단위로,
전체 파이프라인은 실제 API(skipif)로 검증한다."""

import os

import pytest
from langchain_core.messages import HumanMessage
from openpyxl import Workbook

from agent.analyst import (
    AnalystNodes,
    _find_target_file,
    _pick_table_block,
    analyst,
)

FILE = "표본_지출.xlsx"
SHEET = "지출"


@pytest.fixture()
def table_dir(tmp_path, monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "부서별 지출 내역"
    rows = [
        ["부서", "항목", "금액"],
        ["감사1본부", "출장비", 500],
        ["감사1본부", "교육비", 300],
        ["감사2본부", "출장비", 700],
    ]
    for i, row in enumerate(rows, start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(tmp_path / FILE)
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    return tmp_path


def test_find_target_file_attachment_then_mention(table_dir):
    assert _find_target_file([f"[첨부 파일: {FILE}]\n집계해줘"]).name == FILE
    assert _find_target_file([f"{FILE} 부서별 합계"]).name == FILE
    assert _find_target_file(["표본_지출 부서별 합계"]).name == FILE  # stem 언급
    assert _find_target_file(["아무 파일도 언급 안 함"]) is None


def test_find_target_file_fuzzy_tokens(tmp_path, monkeypatch):
    """밑줄·괄호를 생략한 언급을 토큰 겹침으로 잡는다 (겹침 2개 미만은 미매칭)."""
    for name in (
        "데모조서_5400 매출채권 (주)한빛전자 (작성중).xlsx",
        "데모조서_5300 현금및현금성자산 (주)한빛전자.xlsx",
    ):
        wb = Workbook()
        wb.active["A1"] = "x"
        wb.save(tmp_path / name)
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))

    found = _find_target_file(["작성중인 데모조서 5400 매출채권 조서를 검토해줘"])
    assert found is not None and "5400" in found.name

    found = _find_target_file(["데모조서 5300 현금 조서 해석해줘"])
    assert found is not None and "5300" in found.name

    # 토큰 1개 겹침("데모조서")만으로는 특정하지 않는다
    assert _find_target_file(["데모조서 하나 골라줘"]) is None


def test_pick_table_block_prefers_largest(table_dir):
    target = table_dir / FILE
    sheet, ref = _pick_table_block(target, ["부서별 합계 알려줘"])
    assert sheet == SHEET
    assert ref == "A3:C6"  # 제목(A1, 1행짜리)이 아니라 표 본체


def test_conversation_context_builder():
    from langchain_core.messages import AIMessage

    from agent.analyst import _conversation_context

    state = {
        "messages": [
            HumanMessage(content="부서별 합계"),
            AIMessage(content="감사1본부 800, 감사2본부 700"),
            HumanMessage(content="그중 가장 큰 부서만"),
        ]
    }
    ctx = _conversation_context(state)
    assert "사용자: 부서별 합계" in ctx
    assert "어시스턴트: 감사1본부 800" in ctx
    assert "그중 가장 큰 부서만" not in ctx  # 현재 질문은 제외

    # 첫 턴(이력 없음)은 빈 문자열
    assert _conversation_context({"messages": [HumanMessage(content="첫 질문")]}) == ""

    # 긴 메시지는 클립
    long_state = {
        "messages": [HumanMessage(content="가" * 1000), HumanMessage(content="현재")]
    }
    ctx = _conversation_context(long_state)
    assert len(ctx) < 700 and ctx.endswith("…")


def test_triage_fallback_heuristic(table_dir):
    """model=None이면 LLM 분류가 예외 → 파일 언급 유무 휴리스틱으로 폴백."""
    nodes = AnalystNodes(model=None)
    update = nodes.triage(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}] 부서별 합계")]}
    )
    assert update["mode"] == "analysis"
    assert nodes.route_triage(update) == "analysis"

    update = nodes.triage({"messages": [HumanMessage(content="안녕, 뭘 할 수 있어?")]})
    assert update["mode"] == "chat"
    assert nodes.route_triage(update) == "chat"

    update = nodes.triage({"messages": []})  # 빈 입력도 chat으로
    assert update["mode"] == "chat"


def test_explicit_range_with_sheet_overrides_block_pick(table_dir):
    nodes = AnalystNodes(model=None)
    update = nodes.inspect_data(
        {
            "messages": [
                HumanMessage(content=f"[첨부 파일: {FILE}] {SHEET}!A3:C5 범위만 집계해줘")
            ]
        }
    )
    assert update["error"] is None
    assert update["dataset"]["cell_range"] == "A3:C5"  # 자동 선택(A3:C6)이 아니라 명시 범위
    assert update["profile"]["row_count"] == 2


def test_explicit_range_without_sheet_on_single_sheet(table_dir):
    nodes = AnalystNodes(model=None)
    update = nodes.inspect_data(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}] A3:C6 집계해줘")]}
    )
    assert update["error"] is None
    assert update["dataset"] == {"path": FILE, "sheet": SHEET, "cell_range": "A3:C6"}


def test_no_block_error_lists_sheets(table_dir):
    from openpyxl import Workbook as WB

    wb = WB()
    wb.active.title = "표지"
    wb.active["A1"] = "제목뿐"
    wb.save(table_dir / "빈조서.xlsx")
    nodes = AnalystNodes(model=None)
    update = nodes.inspect_data(
        {"messages": [HumanMessage(content="[첨부 파일: 빈조서.xlsx] 집계해줘")]}
    )
    assert update["error"] is not None
    assert "표지" in update["error"]  # 시트 목록 안내
    assert "시트명!A1:C50" in update["error"]  # 범위 지정 안내


def test_inspect_data_without_file_lists_candidates(table_dir):
    nodes = AnalystNodes(model=None)  # inspect는 LLM 미사용
    update = nodes.inspect_data(
        {"messages": [HumanMessage(content="부서별 합계 알려줘")]}
    )
    assert update["error"] and FILE in update["error"]
    assert nodes.route_inspect(update) == "fail"


def test_inspect_data_registers_table(table_dir):
    nodes = AnalystNodes(model=None)
    update = nodes.inspect_data(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}]\n부서별 금액 합계")]}
    )
    assert update["error"] is None
    assert update["dataset"] == {
        "path": FILE,
        "sheet": SHEET,
        "cell_range": "A3:C6",
    }
    assert '"금액"' in update["schema"]
    assert update["profile"]["row_count"] == 3
    assert nodes.route_inspect(update) == "plan"


def test_validate_and_execute_without_llm(table_dir):
    nodes = AnalystNodes(model=None)
    state = nodes.inspect_data(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}] 합계")]}
    )
    state["sql"] = 'SELECT "부서", SUM("금액") AS 합계 FROM data GROUP BY 1 ORDER BY 2 DESC'
    state.update(nodes.validate_sql(state))
    assert state["error"] is None
    state.update(nodes.execute_sql(state))
    assert state["error"] is None
    assert state["result_rows"][0] == ["감사1본부", 800] or state["result_rows"][0] == [
        "감사1본부",
        800.0,
    ]
    assert nodes.route_execution(state) == "answer"

    state.update(nodes.validate_sql({"sql": "DROP TABLE data", "attempts": 2}))
    assert state["error"] and nodes.route_sql(state | {"attempts": 2}) == "fail"


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_followup_uses_conversation_context(table_dir):
    """후속 질의가 이전 대화 맥락('그중')을 SQL로 해석하는지."""
    from langchain_core.messages import AIMessage

    graph = await analyst(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {
            "messages": [
                HumanMessage(content=f"[첨부 파일: {FILE}] 부서별 금액 합계를 알려줘"),
                AIMessage(
                    content=(
                        "부서별 합계: 감사1본부 800, 감사2본부 700 "
                        f"({FILE} [{SHEET}] A3:C6)"
                    )
                ),
                HumanMessage(content="그중 합계가 가장 큰 부서만 다시 보여줘"),
            ]
        }
    )
    answer = result["messages"][-1]
    text = answer.content if isinstance(answer.content, str) else str(answer.content)
    assert "감사1본부" in text  # 맥락('그중')을 해석해 최대 부서를 특정
    assert "오류:" not in text


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_chat_branch_with_real_model(table_dir):
    graph = await analyst(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {"messages": [HumanMessage(content="안녕! 너는 어떤 걸 도와줄 수 있어?")]}
    )
    answer = result["messages"][-1]
    text = answer.content if isinstance(answer.content, str) else str(answer.content)
    assert text.strip()  # 오류 없이 일반 대화 응답
    assert "오류:" not in text
    assert result.get("sql") is None  # SQL 파이프라인을 타지 않았다


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_full_pipeline_with_real_model(table_dir):
    graph = await analyst(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {
            "messages": [
                HumanMessage(
                    content=f"[첨부 파일: {FILE}]\n부서별 금액 합계를 큰 순서로 알려줘"
                )
            ]
        }
    )
    answer = result["messages"][-1]
    text = answer.content if isinstance(answer.content, str) else str(answer.content)
    assert "감사1본부" in text and "800" in text
    assert result["error"] is None
