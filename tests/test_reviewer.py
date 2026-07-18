"""reviewer 그래프 테스트 — 서명란 스캔·증거 수집·보고서 렌더는 단위로,
전체 파이프라인은 실제 API(skipif)로 검증한다."""

import os

import pytest
from langchain_core.messages import HumanMessage
from openpyxl import Workbook

from agent.evidence import collect_workpaper_evidence, scan_signoffs
from agent.reviewer import (
    Finding,
    ReviewFindings,
    ReviewerNodes,
    _render_report,
    reviewer,
)

FILE = "표본_검토조서.xlsx"
SHEET = "5100"


@pytest.fixture()
def review_dir(tmp_path, monkeypatch):
    """서명란(작성자 채움·검토자 공란)과 소형 표·수식을 심은 조서."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "현금 실사 조서"
    ws["A2"] = "작성자"
    ws["B2"] = "김신입"
    ws["A3"] = "검토자"  # B3 공란 — 검토 미완
    rows = [
        ["항목", "장부", "실사"],
        ["소액현금", 100, 100],
        ["보통예금", 900, 900],
    ]
    for i, row in enumerate(rows, start=6):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    ws["B9"] = "=SUM(B7:B8)"
    ws["C9"] = 1000  # 수식 지대 내 하드코딩
    wb.save(tmp_path / FILE)
    monkeypatch.setenv("WORKPAPERS_DIR", str(tmp_path))
    return tmp_path


def test_scan_signoffs_detects_filled_and_blank(review_dir):
    out = scan_signoffs(review_dir / FILE)
    assert f'{SHEET}!A2 "작성자" → 채움(김신입)' in out
    assert f'{SHEET}!A3 "검토자" → 공란' in out


def test_scan_signoffs_ignores_long_sentences(review_dir):
    from openpyxl import load_workbook

    wb = load_workbook(review_dir / FILE)
    wb[SHEET]["A20"] = "검토자는 다음 사항을 확인한 후 서명해야 한다는 안내 문구"
    wb.save(review_dir / FILE)
    out = scan_signoffs(review_dir / FILE)
    assert "A20" not in out


def test_collect_evidence_sections(review_dir):
    out, examined, skipped = collect_workpaper_evidence(FILE)
    assert "워크북:" in out  # overview
    assert "수식 지도:" in out and "하드코딩" in out  # formula_map
    assert "주석·의도 정보:" in out  # annotations
    assert "[서명란 스캔]" in out
    assert "소액현금" in out  # 본문 정독 포함
    # 서명란 스캔은 절단(MAX_EVIDENCE_CHARS)에서 살아남도록 증거 선두에 있다
    assert out.index("[서명란 스캔]") < out.index("수식 지도:")
    assert examined == [SHEET] and skipped == []


def test_collect_evidence_sheet_cap_and_report_scope(review_dir):
    from openpyxl import load_workbook

    wb = load_workbook(review_dir / FILE)
    for i in range(1, 8):  # 기존 1개 + 7개 = 8개 시트 (상한 6개 초과)
        ws = wb.create_sheet(f"추가{i}")
        ws["A1"] = "항목"
        ws["A2"] = "값"
    wb.save(review_dir / FILE)

    out, examined, skipped = collect_workpaper_evidence(FILE)
    assert len(examined) == 6 and examined[0] == SHEET
    assert skipped == ["추가6", "추가7"]
    assert "추가6, 추가7" in out  # 생략 시트가 증거에도 명시

    findings = ReviewFindings(
        workpaper_purpose="p", performed_procedures=[], missing_procedures=[],
        signoff_assessment="s", tieout_findings=[], open_items=[], overall="o",
    )
    report = _render_report(FILE, findings, examined, skipped)
    assert "## 점검 범위" in report
    assert "점검한 시트(6개)" in report
    assert "생략된 시트(2개" in report and "추가7" in report


def test_triage_fallback_heuristic(review_dir):
    """model=None이면 LLM 분류가 예외 → 파일 언급 유무 휴리스틱으로 폴백."""
    nodes = ReviewerNodes(model=None)
    update = nodes.triage(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}] 검토해줘")]}
    )
    assert update["mode"] == "review" and update["error"] is None
    assert nodes.route_triage(update) == "review"

    update = nodes.triage({"messages": [HumanMessage(content="안녕, 사용법 알려줘")]})
    assert update["mode"] == "chat"
    assert nodes.route_triage(update) == "chat"

    update = nodes.triage({"messages": []})  # 빈 입력도 chat으로
    assert update["mode"] == "chat"


def test_locate_without_file_lists_candidates(review_dir):
    nodes = ReviewerNodes(model=None)
    update = nodes.locate({"messages": [HumanMessage(content="조서 검토해줘")]})
    assert update["error"] and FILE in update["error"]
    assert nodes.route_locate(update) == "fail"


def test_render_report_sections_and_severity_order():
    findings = ReviewFindings(
        workpaper_purpose="현금 실사 조서",
        performed_procedures=["실사 대사"],
        missing_procedures=[
            Finding(title="낮음건", severity="낮음", location="5100!A1", detail="d"),
            Finding(
                title="높음건",
                severity="높음",
                location="5100!C9",
                detail="d",
                assertion="완전성",
                risk_if_unresolved="누락된 부채가 계상되지 않은 채 남는다",
            ),
        ],
        signoff_assessment="작성 완료, 검토 서명 공란",
        tieout_findings=[],
        open_items=[],
        overall="검토 서명 전 단계",
    )
    out = _render_report(FILE, findings)
    for section in ("① 조서 개요", "② 수행된 절차", "③ 서명·검토 상태",
                    "④ 검산·tie-out 점검", "⑤ 누락·추가 필요 절차",
                    "⑥ 미결 항목", "⑦ 총평"):
        assert section in out
    assert out.index("높음건") < out.index("낮음건")  # 심각도 정렬
    assert "- (해당 없음)" in out  # 빈 섹션 표기
    assert "기계 수집 증거" in out  # 한계 고지
    # 하위 불릿 구조 — 마크다운에서 줄이 문단으로 접히지 않아야 한다
    assert "\n  - _주장: 완전성 · 미해결 시 위험: 누락된 부채가 계상되지 않은 채 남는다_" in out
    assert "\n  - d" in out  # detail도 제 줄을 갖는다


class _StubInvestigateModel:
    """1라운드에 excel_read_range 호출, 2라운드에 종료하는 스텁."""

    def __init__(self, path: str, sheet: str) -> None:
        from langchain_core.messages import AIMessage as AI

        self._responses = [
            AI(
                content="",
                tool_calls=[
                    {
                        "name": "excel_read_range",
                        "args": {"path": path, "sheet": sheet, "cell_range": "A6:C8"},
                        "id": "t1",
                        "type": "tool_call",
                    }
                ],
            ),
            AI(content="증거 충분"),
        ]

    def bind_tools(self, tools):
        return self

    def invoke(self, messages):
        return self._responses.pop(0)


def test_investigate_with_stub_model(review_dir):
    nodes = ReviewerNodes(model=_StubInvestigateModel(FILE, SHEET))
    update = nodes.investigate(
        {"path": FILE, "question": "점검해줘", "evidence": "(기본 증거)"}
    )
    assert "소액현금" in update["extra_evidence"]  # 도구 결과가 추가 증거로 수집됨
    assert "excel_read_range" in update["extra_evidence"]


def test_investigate_skips_without_model(review_dir):
    nodes = ReviewerNodes(model=None)  # bind_tools 불가 → 보충 없이 통과
    update = nodes.investigate(
        {"path": FILE, "question": "점검해줘", "evidence": "(기본 증거)"}
    )
    assert update["extra_evidence"] == ""


def _findings_with_query() -> ReviewFindings:
    return ReviewFindings(
        workpaper_purpose="현금 실사 조서",
        performed_procedures=[],
        missing_procedures=[
            Finding(
                title="외부조회 미실시",
                severity="높음",
                location="5100!A1",
                detail="d",
                standards_query="외부조회 확인절차",
                source_hint="감사기준",
            )
        ],
        signoff_assessment="s",
        tieout_findings=[],
        open_items=[],
        overall="o",
    )


async def test_cite_attaches_verified_citation(review_dir, monkeypatch):
    from types import SimpleNamespace

    import agent.reviewer as reviewer_mod

    calls = []

    async def search_fn(**kwargs):
        calls.append(("search", kwargs))
        return (
            '{"results": [{"cid": "KSA::505::7", "display": "감사기준서 505 문단 7"}]}'
        )

    async def para_fn(**kwargs):
        calls.append(("get_paragraph", kwargs))
        return (
            '{"paragraphs": [{"cid": "KSA::505::7", '
            '"display": "감사기준서 505 문단 7"}]}'
        )

    async def fake_tools():
        return [
            SimpleNamespace(name="standards_search", coroutine=search_fn),
            SimpleNamespace(name="standards_get_paragraph", coroutine=para_fn),
        ]

    monkeypatch.setattr(reviewer_mod, "get_standards_tools", fake_tools)
    nodes = ReviewerNodes(model=None)  # cite는 LLM 미사용
    update = await nodes.cite({"findings": _findings_with_query().model_dump()})

    result = ReviewFindings.model_validate(update["findings"])
    finding = result.missing_procedures[0]
    assert finding.citation == "감사기준서 505 문단 7"
    assert finding.citation_cid == "KSA::505::7"
    assert calls[0][1]["source_type"] == ["감사기준"]  # source_hint가 필터로 전달
    assert calls[1][0] == "get_paragraph"  # 원문 재확인 수행

    report = _render_report(FILE, result)
    assert "근거: 감사기준서 505 문단 7" in report  # 소견에 병기
    assert "## 근거 목록" in report and "`KSA::505::7`" in report


async def test_cite_runs_findings_concurrently(review_dir, monkeypatch):
    """소견 3건 × 왕복 2회(각 0.2초) — 직렬이면 1.2초, 병렬이면 ~0.4초."""
    import asyncio
    import time
    from types import SimpleNamespace

    import agent.reviewer as reviewer_mod

    async def slow_search(**kwargs):
        await asyncio.sleep(0.2)
        return '{"results": [{"cid": "KSA::505::7", "display": "감사기준서 505 문단 7"}]}'

    async def slow_para(**kwargs):
        await asyncio.sleep(0.2)
        return '{"paragraphs": [{"cid": "KSA::505::7", "display": "감사기준서 505 문단 7"}]}'

    async def fake_tools():
        return [
            SimpleNamespace(name="standards_search", coroutine=slow_search),
            SimpleNamespace(name="standards_get_paragraph", coroutine=slow_para),
        ]

    monkeypatch.setattr(reviewer_mod, "get_standards_tools", fake_tools)
    findings = _findings_with_query()
    findings.missing_procedures *= 3  # 동일 소견 3건
    nodes = ReviewerNodes(model=None)

    start = time.monotonic()
    update = await nodes.cite({"findings": findings.model_dump()})
    elapsed = time.monotonic() - start

    result = ReviewFindings.model_validate(update["findings"])
    assert all(f.citation_cid == "KSA::505::7" for f in result.missing_procedures)
    assert elapsed < 0.9  # 직렬(1.2초)이 아니라 병렬(~0.4초) 실행


class _StubChatModel:
    """1응답에 standards_search 호출, 2응답에 최종 답을 내는 스텁."""

    def __init__(self) -> None:
        from langchain_core.messages import AIMessage as AI

        self._responses = [
            AI(
                content="",
                tool_calls=[
                    {
                        "name": "standards_search",
                        "args": {"query": "외부조회"},
                        "id": "c1",
                        "type": "tool_call",
                    }
                ],
            ),
            AI(content="감사기준서 505 문단 7 (`KSA::505::7`)이 근거입니다."),
        ]

    def bind_tools(self, tools):
        return self

    async def ainvoke(self, messages):
        return self._responses.pop(0)


async def test_chat_uses_standards_tools(review_dir, monkeypatch):
    from types import SimpleNamespace

    import agent.reviewer as reviewer_mod

    calls = []

    async def search_fn(**kwargs):
        calls.append(kwargs)
        return '{"results": [{"cid": "KSA::505::7", "display": "감사기준서 505 문단 7"}]}'

    async def fake_tools():
        return [SimpleNamespace(name="standards_search", coroutine=search_fn)]

    monkeypatch.setattr(reviewer_mod, "get_standards_tools", fake_tools)
    nodes = ReviewerNodes(model=_StubChatModel())
    update = await nodes.chat(
        {"messages": [HumanMessage(content="외부조회 근거 기준이 뭐야?")]}
    )
    assert calls and calls[0]["query"] == "외부조회"  # 도구가 실제 실행됨
    assert "KSA::505::7" in update["messages"][0].content


async def test_chat_degrades_without_mcp(review_dir, monkeypatch):
    from langchain_core.messages import AIMessage as AI

    import agent.reviewer as reviewer_mod

    async def no_tools():
        return []

    class _PlainModel:
        async def ainvoke(self, messages):
            # 도구 없을 때는 bind_tools 없이 곧장 호출된다
            assert "원문 확인 도구가 연결돼 있지 않습니다" in messages[0].content
            return AI(content="사용법 안내입니다.")

    monkeypatch.setattr(reviewer_mod, "get_standards_tools", no_tools)
    nodes = ReviewerNodes(model=_PlainModel())
    update = await nodes.chat({"messages": [HumanMessage(content="안녕")]})
    assert update["messages"][0].content == "사용법 안내입니다."


async def test_cite_graceful_without_mcp(review_dir, monkeypatch):
    import agent.reviewer as reviewer_mod

    async def no_tools():
        return []

    monkeypatch.setattr(reviewer_mod, "get_standards_tools", no_tools)
    nodes = ReviewerNodes(model=None)
    update = await nodes.cite({"findings": _findings_with_query().model_dump()})
    result = ReviewFindings.model_validate(update["findings"])
    assert result.missing_procedures[0].citation == ""  # 인용 없이 통과
    assert "## 근거 목록" not in _render_report(FILE, result)


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_chat_branch_with_real_model(review_dir):
    graph = await reviewer(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {"messages": [HumanMessage(content="안녕! 너는 어떤 검토를 해줘?")]}
    )
    answer = result["messages"][-1]
    text = answer.content if isinstance(answer.content, str) else str(answer.content)
    assert text.strip() and "오류:" not in text
    assert "# 조서 검토 보고" not in text  # 검토 파이프라인을 타지 않았다


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"), reason="실 API 키 필요"
)
async def test_full_pipeline_with_real_model(review_dir):
    graph = await reviewer(
        {"configurable": {"model": "anthropic:claude-haiku-4-5-20251001"}}
    )
    result = await graph.ainvoke(
        {"messages": [HumanMessage(content=f"[첨부 파일: {FILE}]\n이 조서 완성도 점검해줘")]}
    )
    text = result["messages"][-1].content
    assert "# 조서 검토 보고" in text and "⑦ 총평" in text
    assert result["error"] is None
